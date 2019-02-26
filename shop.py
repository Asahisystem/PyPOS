import sys
import sip
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
import datetime
import openpyxl
import os
import pyqtgraph as pg
import webbrowser

allplice = int(0)
paymanay = str("")
change = int(0)
analizepath = os.path.abspath('POSanalizer.xlsx')
setpath = os.path.abspath('Setting.xlsx')
zcode = int(0)
info = int(0)
jcode = int(0)
p1tag = int(0)
p2tag = int(0)
p3tag = int(0)
p4tag = int(0)
p5tag = int(0)
p6tag = int(0)
p7tag = int(0)
p8tag = int(0)
p9tag = int(0)
p10tag = int(0)
p11tag = int(0)
p12tag = int(0)
p13tag = int(0)
p14tag = int(0)
p15tag = int(0)
p16tag = int(0)
p17tag = int(0)
p18tag = int(0)
p19tag = int(0)
p20tag = int(0)
p21tag = int(0)
p22tag = int(0)
p23tag = int(0)
p24tag = int(0)
p25tag = int(0)
p26tag = int(0)
p27tag = int(0)
p28tag = int(0)
p29tag = int(0)
p30tag = int(0)
p31tag = int(0)
p32tag = int(0)
p33tag = int(0)
p34tag = int(0)
p35tag = int(0)
p36tag = int(0)

nametag1 = str("")
nametag2 = str("")
nametag3 = str("")
nametag4 = str("")
nametag5 = str("")
nametag6 = str("")
nametag7 = str("")
nametag8 = str("")
nametag9 = str("")
nametag10 = str("")
nametag11 = str("")
nametag12 = str("")
nametag13 = str("")
nametag14 = str("")
nametag15 = str("")
nametag16 = str("")
nametag17 = str("")
nametag18 = str("")
nametag19 = str("")
nametag20 = str("")
nametag21 = str("")
nametag22 = str("")
nametag23 = str("")
nametag24 = str("")
nametag25 = str("")
nametag26 = str("")
nametag27 = str("")
nametag28 = str("")
nametag29 = str("")
nametag30 = str("")
nametag31 = str("")
nametag32 = str("")
nametag33 = str("")
nametag34 = str("")
nametag35 = str("")
nametag36 = str("")

sp1tag = int(0)
sp2tag = int(0)
sp3tag = int(0)
sp4tag = int(0)
sp5tag = int(0)
sp6tag = int(0)
sp7tag = int(0)
sp8tag = int(0)
sp9tag = int(0)
sp10tag = int(0)
sp11tag = int(0)
sp12tag = int(0)
sp13tag = int(0)
sp14tag = int(0)
sp15tag = int(0)
sp16tag = int(0)
sp17tag = int(0)
sp18tag = int(0)
sp19tag = int(0)
sp20tag = int(0)
sp21tag = int(0)
sp22tag = int(0)
sp23tag = int(0)
sp24tag = int(0)
sp25tag = int(0)
sp26tag = int(0)
sp27tag = int(0)
sp28tag = int(0)
sp29tag = int(0)
sp30tag = int(0)
sp31tag = int(0)
sp32tag = int(0)
sp33tag = int(0)
sp34tag = int(0)
sp35tag = int(0)
sp36tag = int(0)

snametag1 = str("")
snametag2 = str("")
snametag3 = str("")
snametag4 = str("")
snametag5 = str("")
snametag6 = str("")
snametag7 = str("")
snametag8 = str("")
snametag9 = str("")
snametag10 = str("")
snametag11 = str("")
snametag12 = str("")
snametag13 = str("")
snametag14 = str("")
snametag15 = str("")
snametag16 = str("")
snametag17 = str("")
snametag18 = str("")
snametag19 = str("")
snametag20 = str("")
snametag21 = str("")
snametag22 = str("")
snametag23 = str("")
snametag24 = str("")
snametag25 = str("")
snametag26 = str("")
snametag27 = str("")
snametag28 = str("")
snametag29 = str("")
snametag30 = str("")
snametag31 = str("")
snametag32 = str("")
snametag33 = str("")
snametag34 = str("")
snametag35 = str("")
snametag36 = str("")

dp2tag = int(0)
dp3tag = int(0)
dp4tag = int(0)
dp5tag = int(0)
dp6tag = int(0)
dp7tag = int(0)
dp8tag = int(0)
dp9tag = int(0)
dp10tag = int(0)
dp11tag = int(0)
dp12tag = int(0)
dp13tag = int(0)
dp14tag = int(0)
dp15tag = int(0)
dp16tag = int(0)
dp17tag = int(0)
dp18tag = int(0)
dp19tag = int(0)
dp20tag = int(0)
dp21tag = int(0)
dp22tag = int(0)
dp23tag = int(0)
dp24tag = int(0)
dp25tag = int(0)
dp26tag = int(0)
dp27tag = int(0)
dp28tag = int(0)
dp29tag = int(0)
dp30tag = int(0)
dp31tag = int(0)
dp32tag = int(0)
dp33tag = int(0)
dp34tag = int(0)
dp35tag = int(0)
dp36tag = int(0)
dnametag1 = str("")
dnametag2 = str("")
dnametag3 = str("")
dnametag4 = str("")
dnametag5 = str("")
dnametag6 = str("")
dnametag7 = str("")
dnametag8 = str("")
dnametag9 = str("")
dnametag10 = str("")
dnametag11 = str("")
dnametag12 = str("")
dnametag13 = str("")
dnametag14 = str("")
dnametag15 = str("")
dnametag16 = str("")
dnametag17 = str("")
dnametag18 = str("")
dnametag19 = str("")
dnametag20 = str("")
dnametag21 = str("")
dnametag22 = str("")
dnametag23 = str("")
dnametag24 = str("")
dnametag25 = str("")
dnametag26 = str("")
dnametag27 = str("")
dnametag28 = str("")
dnametag29 = str("")
dnametag30 = str("")
dnametag31 = str("")
dnametag32 = str("")
dnametag33 = str("")
dnametag34 = str("")
dnametag35 = str("")
dnametag36 = str("")
nd1 = int(0)
nd2 = int(0)
nd3 = int(0)
nd4 = int(0)
nd5 = int(0)
nd6 = int(0)
nd7 = int(0)
nd8 = int(0)
nd9 = int(0)
nd10 = int(0)
nd11 = int(0)
nd12 = int(0)
nd13 = int(0)
nd14 = int(0)
nd15 = int(0)
nd16 = int(0)
nd17 = int(0)
nd18 = int(0)
nd19 = int(0)
nd20 = int(0)
nd21 = int(0)
nd22 = int(0)
nd23 = int(0)
nd24 = int(0)
nd25 = int(0)
nd26 = int(0)
nd27 = int(0)
nd28 = int(0)
nd29 = int(0)
nd30 = int(0)
nd31 = int(0)
nd32 = int(0)
nd33 = int(0)
nd34 = int(0)
nd35 = int(0)
nd36 = int(0)







class Model(QStandardItemModel):
    def __init__(self, row, column, parent=None):
        super(Model, self).__init__(row, column, parent)
        self.row = row
        self.column = column



    def additeminrow(self, texts):
        self.insertRows(0, 1)
        for c in range(self.column):
            self.setData(self.index(0,c), texts[c])
            self.row +=1







class Tab1Widget(QWidget):
    def __init__(self, parent=None):
        super(Tab1Widget, self).__init__()

        Tab1Widget.name1 = QPushButton('indname1', self)#//Button
        Tab1Widget.name1.clicked.connect(self.n1fun)
        Tab1Widget.name1.setGeometry(0,0,100,50)
        Tab1Widget.name2 = QPushButton('indname2', self)
        Tab1Widget.name2.clicked.connect(self.n2fun)
        Tab1Widget.name2.setGeometry(0,50,100,50)
        Tab1Widget.name3 = QPushButton('indname3', self)
        Tab1Widget.name3.clicked.connect(self.n3fun)
        Tab1Widget.name3.setGeometry(0,100,100,50)
        Tab1Widget.name4 = QPushButton('indname4', self)
        Tab1Widget.name4.clicked.connect(self.n4fun)
        Tab1Widget.name4.setGeometry(0,150,100,50)
        Tab1Widget.name5 = QPushButton('indnam5', self)
        Tab1Widget.name5.clicked.connect(self.n5fun)
        Tab1Widget.name5.setGeometry(0,200,100,50)
        Tab1Widget.name6 = QPushButton('indname6', self)
        Tab1Widget.name6.clicked.connect(self.n6fun)
        Tab1Widget.name6.setGeometry(0,250,100,50)
        Tab1Widget.name7 = QPushButton('indname7', self)
        Tab1Widget.name7.clicked.connect(self.n7fun)
        Tab1Widget.name7.setGeometry(0,300,100,50)
        Tab1Widget.name8 = QPushButton('indname8', self)
        Tab1Widget.name8.clicked.connect(self.n8fun)
        Tab1Widget.name8.setGeometry(0,350,100,50)
        Tab1Widget.name9 = QPushButton('indname9', self)
        Tab1Widget.name9.clicked.connect(self.n9fun)
        Tab1Widget.name9.setGeometry(0,400,100,50)

        Tab1Widget.name10 = QPushButton('indname10', self)
        Tab1Widget.name10.clicked.connect(self.n10fun)
        Tab1Widget.name10.setGeometry(100,0,100,50)
        Tab1Widget.name11 = QPushButton('indname11', self)
        Tab1Widget.name11.clicked.connect(self.n11fun)
        Tab1Widget.name11.setGeometry(100,50,100,50)
        Tab1Widget.name12 = QPushButton('indname12', self)
        Tab1Widget.name12.clicked.connect(self.n12fun)
        Tab1Widget.name12.setGeometry(100,100,100,50)
        Tab1Widget.name13 = QPushButton('indname13', self)
        Tab1Widget.name13.clicked.connect(self.n13fun)
        Tab1Widget.name13.setGeometry(100,150,100,50)
        Tab1Widget.name14 = QPushButton('indnam14', self)
        Tab1Widget.name14.clicked.connect(self.n14fun)
        Tab1Widget.name14.setGeometry(100,200,100,50)
        Tab1Widget.name15 = QPushButton('indname15', self)
        Tab1Widget.name15.clicked.connect(self.n15fun)
        Tab1Widget.name15.setGeometry(100,250,100,50)
        Tab1Widget.name16= QPushButton('indname16', self)
        Tab1Widget.name16.clicked.connect(self.n16fun)
        Tab1Widget.name16.setGeometry(100,300,100,50)
        Tab1Widget.name17 = QPushButton('indname17', self)
        Tab1Widget.name17.clicked.connect(self.n17fun)
        Tab1Widget.name17.setGeometry(100,350,100,50)
        Tab1Widget.name18 = QPushButton('indname18', self)
        Tab1Widget.name18.clicked.connect(self.n18fun)
        Tab1Widget.name18.setGeometry(100,400,100,50)

        Tab1Widget.name19 = QPushButton('indname1', self)#//Button
        Tab1Widget.name19.clicked.connect(self.n19fun)
        Tab1Widget.name19.setGeometry(200,0,100,50)
        Tab1Widget.name20 = QPushButton('indname2', self)
        Tab1Widget.name20.clicked.connect(self.n20fun)
        Tab1Widget.name20.setGeometry(200,50,100,50)
        Tab1Widget.name21 = QPushButton('indname3', self)
        Tab1Widget.name21.clicked.connect(self.n21fun)
        Tab1Widget.name21.setGeometry(200,100,100,50)
        Tab1Widget.name22 = QPushButton('indname4', self)
        Tab1Widget.name22.clicked.connect(self.n22fun)
        Tab1Widget.name22.setGeometry(200,150,100,50)
        Tab1Widget.name23 = QPushButton('indnam5', self)
        Tab1Widget.name23.clicked.connect(self.n23fun)
        Tab1Widget.name23.setGeometry(200,200,100,50)
        Tab1Widget.name24 = QPushButton('indname6', self)
        Tab1Widget.name24.clicked.connect(self.n24fun)
        Tab1Widget.name24.setGeometry(200,250,100,50)
        Tab1Widget.name25 = QPushButton('indname7', self)
        Tab1Widget.name25.clicked.connect(self.n25fun)
        Tab1Widget.name25.setGeometry(200,300,100,50)
        Tab1Widget.name26 = QPushButton('indname8', self)
        Tab1Widget.name26.clicked.connect(self.n26fun)
        Tab1Widget.name26.setGeometry(200,350,100,50)
        Tab1Widget.name27 = QPushButton('indname9', self)
        Tab1Widget.name27.clicked.connect(self.n27fun)
        Tab1Widget.name27.setGeometry(200,400,100,50)

        Tab1Widget.name28 = QPushButton('indname10', self)
        Tab1Widget.name28.clicked.connect(self.n28fun)
        Tab1Widget.name28.setGeometry(300,0,100,50)
        Tab1Widget.name29 = QPushButton('indname11', self)
        Tab1Widget.name29.clicked.connect(self.n29fun)
        Tab1Widget.name29.setGeometry(300,50,100,50)
        Tab1Widget.name30 = QPushButton('indname12', self)
        Tab1Widget.name30.clicked.connect(self.n30fun)
        Tab1Widget.name30.setGeometry(300,100,100,50)
        Tab1Widget.name31 = QPushButton('indname13', self)
        Tab1Widget.name31.clicked.connect(self.n31fun)
        Tab1Widget.name31.setGeometry(300,150,100,50)
        Tab1Widget.name32 = QPushButton('indnam14', self)
        Tab1Widget.name32.clicked.connect(self.n32fun)
        Tab1Widget.name32.setGeometry(300,200,100,50)
        Tab1Widget.name33 = QPushButton('indname15', self)
        Tab1Widget.name33.clicked.connect(self.n33fun)
        Tab1Widget.name33.setGeometry(300,250,100,50)
        Tab1Widget.name34= QPushButton('indname16', self)
        Tab1Widget.name34.clicked.connect(self.n34fun)
        Tab1Widget.name34.setGeometry(300,300,100,50)
        Tab1Widget.name35 = QPushButton('indname17', self)
        Tab1Widget.name35.clicked.connect(self.n35fun)
        Tab1Widget.name35.setGeometry(300,350,100,50)
        Tab1Widget.name36 = QPushButton('indname18', self)
        Tab1Widget.name36.clicked.connect(self.n36fun)
        Tab1Widget.name36.setGeometry(300,400,100,50)

        self.setStyleSheet('QPushButton{color:#5f5fff; background-color: #ffffff; padding: 0.3px 1px; margin: 2px 1px; font-weight: bold; border: solid; border-width: 1px; border-color: #5f5fff; border-radius: 3px}' ' QPushButton:hover {color: #fff; background-color: #5f5fff;}')

    def n1fun(self):
        global allplice
        global p1tag
        global nametag1
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E3"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E3"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p1tag))
            Tab1Widget.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag1)
            one = str(p1tag)
            display_list = [str(time), string, one]
            Tab1Widget.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E3"].value
            aimpoint = analize["D3"].value
            analize["C3"].value = nametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p1tag))
            analize["E3"].value = inputcode
            analize["D3"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag1)+":"+"￥"+str(p1tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag1+'が残り少なくなっています')







        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n2fun(self):
        global allplice
        global p2tag
        global nametag2
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E4"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E4"].value = zaiko1
            tag.save(setpath)


            allplice = (int(allplice)+int(p2tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag2)
            one = str(p2tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E4"].value
            aimpoint = analize["D4"].value
            analize["C4"].value = nametag2
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p2tag))
            analize["E4"].value = inputcode
            analize["D4"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag2)+":"+"￥"+str(p2tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag2+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n3fun(self):
        global allplice
        global p3tag
        global nametag3
        global setpath
        global analizepath
        global info


        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E5"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E5"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p3tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag3)
            one = str(p3tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E5"].value
            aimpoint = analize["D5"].value
            analize["C5"].value = nametag3
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p3tag))
            analize["E5"].value = inputcode
            analize["D5"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag3)+":"+"￥"+str(p3tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag3+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n4fun(self):
        global allplice
        global p4tag
        global nametag4
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E6"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E6"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p4tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag4)
            one = str(p4tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E6"].value
            aimpoint = analize["D6"].value
            analize["C6"].value = nametag4
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p4tag))
            analize["E6"].value = inputcode
            analize["D6"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag4)+":"+"￥"+str(p4tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag41+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n5fun(self):
        global allplice
        global p5tag
        global nametag5
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E7"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E7"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p5tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag5)
            one = str(p5tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E7"].value
            aimpoint = analize["D7"].value
            analize["C7"].value = nametag5
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p5tag))
            analize["E7"].value = inputcode
            analize["D7"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag5)+":"+"￥"+str(p5tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag5+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n6fun(self):
        global allplice
        global p6tag
        global nametag6
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E8"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E8"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p6tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag6)
            one = str(p6tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E8"].value
            aimpoint = analize["D8"].value
            analize["C8"].value = nametag6
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p6tag))
            analize["E8"].value = inputcode
            analize["D8"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag6)+":"+"￥"+str(p6tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag6+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')

    def n7fun(self):
        global allplice
        global p7tag
        global nametag7
        global setpath
        global info
        global analizepath

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E9"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E9"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p7tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag7)
            one = str(p7tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E9"].value
            aimpoint = analize["D9"].value
            analize["C9"].value = nametag7
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p7tag))
            analize["E9"].value = inputcode
            analize["D9"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag7)+":"+"￥"+str(p7tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag7+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n8fun(self):
        global allplice
        global p8tag
        global nametag8
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E10"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E10"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p8tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag8)
            one = str(p8tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E10"].value
            aimpoint = analize["D10"].value
            analize["C10"].value = nametag8
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p8tag))
            analize["E10"].value = inputcode
            analize["D10"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag8)+":"+"￥"+str(p8tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag8+'が残り少なくなっています')


        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')

    def n9fun(self):
        global allplice
        global p9tag
        global nametag9
        global info
        global analizepath
        global setpath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E11"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E11"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p9tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag9)
            one = str(p9tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E11"].value
            aimpoint = analize["D11"].value
            analize["C11"].value = nametag9
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p9tag))
            analize["E11"].value = inputcode
            analize["D11"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag9)+":"+"￥"+str(p9tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag9+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n10fun(self):
        global allplice
        global p10tag
        global nametag10
        global analizepath
        global setpath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E12"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E12"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p10tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag10)
            one = str(p10tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E12"].value
            aimpoint = analize["D12"].value
            analize["C12"].value = nametag10
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p10tag))
            analize["E12"].value = inputcode
            analize["D12"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag10)+":"+"￥"+str(p10tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag10+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n11fun(self):
        global allplice
        global p11tag
        global nametag11
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E13"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E13"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p11tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag11)
            one = str(p11tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E13"].value
            aimpoint = analize["D13"].value
            analize["C13"].value = nametag11
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p11tag))
            analize["E13"].value = inputcode
            analize["D13"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag11)+":"+"￥"+str(p11tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag11+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n12fun(self):
        global allplice
        global info
        global p12tag
        global nametag12
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E14"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E14"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p12tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag12)
            one = str(p12tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E14"].value
            aimpoint = analize["D14"].value
            analize["C14"].value = nametag12
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p12tag))
            analize["E14"].value = inputcode
            analize["D14"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag12)+":"+"￥"+str(p12tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag12+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n13fun(self):
        global p13tag
        global allplice
        global nametag13
        global info
        global analizepath
        global setpath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E15"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E15"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p13tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag13)
            one = str(p13tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E15"].value
            aimpoint = analize["D15"].value
            analize["C15"].value = nametag13
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p13tag))
            analize["E15"].value = inputcode
            analize["D15"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag13)+":"+"￥"+str(p13tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag13+'が残り少なくなっています')


        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')

    def n14fun(self):
        global allplice
        global p14tag
        global nametag14
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E16"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E16"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p14tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag14)
            one = str(p14tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E16"].value
            aimpoint = analize["D16"].value
            analize["C16"].value = nametag14
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p14tag))
            analize["E16"].value = inputcode
            analize["D16"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag14)+":"+"￥"+str(p14tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag14+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n15fun(self):
        global allplice
        global p15tag
        global nametag15
        global analizepath
        global setpath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E17"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E17"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p15tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag15)
            one = str(p15tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E17"].value
            aimpoint = analize["D17"].value
            analize["C17"].value = nametag15
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p15tag))
            analize["E17"].value = inputcode
            analize["D17"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag15)+":"+"￥"+str(p15tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag15+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n16fun(self):
        global allplice
        global p16tag
        global nametag16
        global analizepath
        global setpath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E18"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E18"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p16tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag16)
            one = str(p16tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E18"].value
            aimpoint = analize["D18"].value
            analize["C18"].value = nametag16
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p16tag))
            analize["E18"].value = inputcode
            analize["D18"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag16)+":"+"￥"+str(p16tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag16+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n17fun(self):
        global allplice
        global p17tag
        global nametag17
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E19"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E19"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p17tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag17)
            one = str(p17tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E19"].value
            aimpoint = analize["D19"].value
            analize["C19"].value = nametag17
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p17tag))
            analize["E19"].value = inputcode
            analize["D19"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag17)+":"+"￥"+str(p17tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag17+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n18fun(self):
        global allplice
        global p18tag
        global nametag18
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E20"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E20"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p18tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag18)
            one = str(p18tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E20"].value
            aimpoint = analize["D20"].value
            analize["C20"].value = nametag18
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p18tag))
            analize["E20"].value = inputcode
            analize["D20"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag18)+":"+"￥"+str(p18tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag18+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n19fun(self):
        global allplice
        global p19tag
        global nametag19
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k3"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k3"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p19tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag19)
            one = str(p19tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E21"].value
            aimpoint = analize["D21"].value
            analize["C21"].value = nametag19
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p19tag))
            analize["E21"].value = inputcode
            analize["D21"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag19)+":"+"￥"+str(p19tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag19+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n20fun(self):
        global allplice
        global p20tag
        global nametag20
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k4"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k4"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p20tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag20)
            one = str(p20tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E22"].value
            aimpoint = analize["D22"].value
            analize["C22"].value = nametag20
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p20tag))
            analize["E22"].value = inputcode
            analize["D22"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag20)+":"+"￥"+str(p20tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag20+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n21fun(self):
        global allplice
        global p21tag
        global nametag21
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k5"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k5"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p21tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag21)
            one = str(p21tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E23"].value
            aimpoint = analize["D23"].value
            analize["C23"].value = nametag21
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p21tag))
            analize["E23"].value = inputcode
            analize["D23"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag21)+":"+"￥"+str(p21tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag21+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n22fun(self):
        global allplice
        global p22tag
        global nametag22
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k6"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k6"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p22tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag22)
            one = str(p22tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E24"].value
            aimpoint = analize["D24"].value
            analize["C24"].value = nametag22
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p22tag))
            analize["E24"].value = inputcode
            analize["D24"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag22)+":"+"￥"+str(p22tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag22+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n23fun(self):
        global allplice
        global p23tag
        global nametag23
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k7"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k7"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p23tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag23)
            one = str(p23tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E25"].value
            aimpoint = analize["D25"].value
            analize["C25"].value = nametag23
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p23tag))
            analize["E25"].value = inputcode
            analize["D25"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag23)+":"+"￥"+str(p23tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag23+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n24fun(self):
        global allplice
        global p24tag
        global nametag24
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k8"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k8"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p24tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag24)
            one = str(p24tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E26"].value
            aimpoint = analize["D26"].value
            analize["C3"].value = nametag24
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p24tag))
            analize["E26"].value = inputcode
            analize["D26"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag24)+":"+"￥"+str(p24tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag24+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n25fun(self):
        global allplice
        global p25tag
        global nametag25
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k9"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k9"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p25tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag25)
            one = str(p25tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E27"].value
            aimpoint = analize["D27"].value
            analize["C27"].value = nametag25
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p25tag))
            analize["E27"].value = inputcode
            analize["D27"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag25)+":"+"￥"+str(p25tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag25+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n26fun(self):
        global allplice
        global p26tag
        global nametag26
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k10"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k10"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p26tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag26)
            one = str(p26tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E28"].value
            aimpoint = analize["D28"].value
            analize["C28"].value = nametag26
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p26tag))
            analize["E28"].value = inputcode
            analize["D28"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag26)+":"+"￥"+str(p26tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag26+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n27fun(self):
        global allplice
        global p27tag
        global nametag27
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k11"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k11"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p27tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag27)
            one = str(p27tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E29"].value
            aimpoint = analize["D29"].value
            analize["C29"].value = nametag27
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p27tag))
            analize["E29"].value = inputcode
            analize["D29"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag27)+":"+"￥"+str(p27tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag27+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n28fun(self):
        global allplice
        global p28tag
        global nametag28
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k12"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k12"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p28tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag28)
            one = str(p28tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E30"].value
            aimpoint = analize["D30"].value
            analize["C30"].value = nametag28
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p28tag))
            analize["E30"].value = inputcode
            analize["D30"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag28)+":"+"￥"+str(p28tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag28+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n29fun(self):
        global allplice
        global p29tag
        global nametag29
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k13"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k13"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p29tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag29)
            one = str(p29tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E31"].value
            aimpoint = analize["D31"].value
            analize["C31"].value = nametag29
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p29tag))
            analize["E31"].value = inputcode
            analize["D31"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag29)+":"+"￥"+str(p29tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag29+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n30fun(self):
        global allplice
        global p30tag
        global nametag30
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k14"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k14"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p30tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag30)
            one = str(p30tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E32"].value
            aimpoint = analize["D32"].value
            analize["C32"].value = nametag30
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p30tag))
            analize["E32"].value = inputcode
            analize["D32"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag30)+":"+"￥"+str(p30tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag30+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n31fun(self):
        global allplice
        global analizepath
        global p31tag
        global info
        global nametag31
        global setpath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k15"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k15"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p31tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag31)
            one = str(p31tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E33"].value
            aimpoint = analize["D33"].value
            analize["C33"].value = nametag31
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p31tag))
            analize["E33"].value = inputcode
            analize["D33"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag31)+":"+"￥"+str(p31tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag31+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n32fun(self):
        global allplice
        global p32tag
        global info
        global nametag32
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k16"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k16"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p32tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag32)
            one = str(p32tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E34"].value
            aimpoint = analize["D34"].value
            analize["C34"].value = nametag32
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p32tag))
            analize["E34"].value = inputcode
            analize["D34"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag32)+":"+"￥"+str(p32tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag32+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n33fun(self):
        global allplice
        global p33tag
        global nametag33
        global info
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k17"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k17"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p33tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag33)
            one = str(p33tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E35"].value
            aimpoint = analize["D35"].value
            analize["C35"].value = nametag33
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p33tag))
            analize["E35"].value = inputcode
            analize["D35"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag33)+":"+"￥"+str(p33tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag33+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n34fun(self):
        global allplice
        global p34tag
        global nametag34
        global info
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k18"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k18"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p34tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag34)
            one = str(p34tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E36"].value
            aimpoint = analize["D36"].value
            analize["C36"].value = nametag34
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p34tag))
            analize["E36"].value = inputcode
            analize["D36"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag34)+":"+"￥"+str(p34tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag34+'が残り少なくなっています')


        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n35fun(self):
        global allplice
        global p35tag
        global nametag35
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k19"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k19"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(p35tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag35)
            one = str(p35tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E37"].value
            aimpoint = analize["D37"].value
            analize["C37"].value = nametag35
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p35tag))
            analize["E37"].value = inputcode
            analize["D37"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag35)+":"+"￥"+str(p35tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag35+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n36fun(self):
        global allplice
        global p36tag
        global info
        global nametag36
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k20"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k20"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p36tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag36)
            one = str(p36tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["E38"].value
            aimpoint = analize["D38"].value
            analize["C38"].value = nametag36
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p36tag))
            analize["E38"].value = inputcode
            analize["D38"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag36)+":"+"￥"+str(p36tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag36+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




class Tab2Widget(QWidget):
    def __init__(self, parent=None):
            super(Tab2Widget, self).__init__()

            Tab1Widget.sname1 = QPushButton('indname1', self)
            Tab1Widget.sname1.clicked.connect(self.n1fun)
            Tab1Widget.sname1.setGeometry(0,0,100,50)
            Tab1Widget.sname2 = QPushButton('indname2', self)
            Tab1Widget.sname2.clicked.connect(self.n2fun)
            Tab1Widget.sname2.setGeometry(0,50,100,50)
            Tab1Widget.sname3 = QPushButton('indname3', self)
            Tab1Widget.sname3.clicked.connect(self.n3fun)
            Tab1Widget.sname3.setGeometry(0,100,100,50)
            Tab1Widget.sname4 = QPushButton('indname4', self)
            Tab1Widget.sname4.clicked.connect(self.n4fun)
            Tab1Widget.sname4.setGeometry(0,150,100,50)
            Tab1Widget.sname5 = QPushButton('indnam5', self)
            Tab1Widget.sname5.clicked.connect(self.n5fun)
            Tab1Widget.sname5.setGeometry(0,200,100,50)
            Tab1Widget.sname6 = QPushButton('indname6', self)
            Tab1Widget.sname6.clicked.connect(self.n6fun)
            Tab1Widget.sname6.setGeometry(0,250,100,50)
            Tab1Widget.sname7 = QPushButton('indname7', self)
            Tab1Widget.sname7.clicked.connect(self.n7fun)
            Tab1Widget.sname7.setGeometry(0,300,100,50)
            Tab1Widget.sname8 = QPushButton('indname8', self)
            Tab1Widget.sname8.clicked.connect(self.n8fun)
            Tab1Widget.sname8.setGeometry(0,350,100,50)
            Tab1Widget.sname9 = QPushButton('indname9', self)
            Tab1Widget.sname9.clicked.connect(self.n9fun)
            Tab1Widget.sname9.setGeometry(0,400,100,50)

            Tab1Widget.sname10 = QPushButton('indname10', self)
            Tab1Widget.sname10.clicked.connect(self.n10fun)
            Tab1Widget.sname10.setGeometry(100,0,100,50)
            Tab1Widget.sname11 = QPushButton('indname11', self)
            Tab1Widget.sname11.clicked.connect(self.n11fun)
            Tab1Widget.sname11.setGeometry(100,50,100,50)
            Tab1Widget.sname12 = QPushButton('indname12', self)
            Tab1Widget.sname12.clicked.connect(self.n12fun)
            Tab1Widget.sname12.setGeometry(100,100,100,50)
            Tab1Widget.sname13 = QPushButton('indname13', self)
            Tab1Widget.sname13.clicked.connect(self.n13fun)
            Tab1Widget.sname13.setGeometry(100,150,100,50)
            Tab1Widget.sname14 = QPushButton('indnam14', self)
            Tab1Widget.sname14.clicked.connect(self.n14fun)
            Tab1Widget.sname14.setGeometry(100,200,100,50)
            Tab1Widget.sname15 = QPushButton('indname15', self)
            Tab1Widget.sname15.clicked.connect(self.n15fun)
            Tab1Widget.sname15.setGeometry(100,250,100,50)
            Tab1Widget.sname16= QPushButton('indname16', self)
            Tab1Widget.sname16.clicked.connect(self.n16fun)
            Tab1Widget.sname16.setGeometry(100,300,100,50)
            Tab1Widget.sname17 = QPushButton('indname17', self)
            Tab1Widget.sname17.clicked.connect(self.n17fun)
            Tab1Widget.sname17.setGeometry(100,350,100,50)
            Tab1Widget.sname18 = QPushButton('indname18', self)
            Tab1Widget.sname18.clicked.connect(self.n18fun)
            Tab1Widget.sname18.setGeometry(100,400,100,50)
            Tab1Widget.sname19 = QPushButton('indname1', self)
            Tab1Widget.sname19.clicked.connect(self.n19fun)
            Tab1Widget.sname20 = QPushButton('indname2', self)
            Tab1Widget.sname20.clicked.connect(self.n20fun)
            Tab1Widget.sname20.setGeometry(200,50,100,50)
            Tab1Widget.sname21 = QPushButton('indname3', self)
            Tab1Widget.sname21.clicked.connect(self.n21fun)
            Tab1Widget.sname21.setGeometry(200,100,100,50)
            Tab1Widget.sname22 = QPushButton('indname4', self)
            Tab1Widget.sname22.clicked.connect(self.n22fun)
            Tab1Widget.sname22.setGeometry(200,150,100,50)
            Tab1Widget.sname23 = QPushButton('indnam5', self)
            Tab1Widget.sname23.clicked.connect(self.n23fun)
            Tab1Widget.sname23.setGeometry(200,200,100,50)
            Tab1Widget.sname24 = QPushButton('indname6', self)
            Tab1Widget.sname24.clicked.connect(self.n24fun)
            Tab1Widget.sname24.setGeometry(200,250,100,50)
            Tab1Widget.sname25 = QPushButton('indname7', self)
            Tab1Widget.sname25.clicked.connect(self.n25fun)
            Tab1Widget.sname25.setGeometry(200,300,100,50)
            Tab1Widget.sname26 = QPushButton('indname8', self)
            Tab1Widget.sname26.clicked.connect(self.n26fun)
            Tab1Widget.sname26.setGeometry(200,350,100,50)
            Tab1Widget.sname27 = QPushButton('indname9', self)
            Tab1Widget.sname19.setGeometry(200,0,100,50)
            Tab1Widget.sname27.clicked.connect(self.n27fun)
            Tab1Widget.sname27.setGeometry(200,400,100,50)

            Tab1Widget.sname28 = QPushButton('indname10', self)
            Tab1Widget.sname28.clicked.connect(self.n28fun)
            Tab1Widget.sname28.setGeometry(300,0,100,50)
            Tab1Widget.sname29 = QPushButton('indname11', self)
            Tab1Widget.sname29.clicked.connect(self.n29fun)
            Tab1Widget.sname29.setGeometry(300,50,100,50)
            Tab1Widget.sname30 = QPushButton('indname12', self)
            Tab1Widget.sname30.clicked.connect(self.n30fun)
            Tab1Widget.sname30.setGeometry(300,100,100,50)
            Tab1Widget.sname31 = QPushButton('indname13', self)
            Tab1Widget.sname31.clicked.connect(self.n31fun)
            Tab1Widget.sname31.setGeometry(300,150,100,50)
            Tab1Widget.sname32 = QPushButton('indnam14', self)
            Tab1Widget.sname32.clicked.connect(self.n32fun)
            Tab1Widget.sname32.setGeometry(300,200,100,50)
            Tab1Widget.sname33 = QPushButton('indname15', self)
            Tab1Widget.sname33.clicked.connect(self.n33fun)
            Tab1Widget.sname33.setGeometry(300,250,100,50)
            Tab1Widget.sname34= QPushButton('indname16', self)
            Tab1Widget.sname34.clicked.connect(self.n34fun)
            Tab1Widget.sname34.setGeometry(300,300,100,50)
            Tab1Widget.sname35 = QPushButton('indname17', self)
            Tab1Widget.sname35.clicked.connect(self.n35fun)
            Tab1Widget.sname35.setGeometry(300,350,100,50)
            Tab1Widget.sname36 = QPushButton('indname18', self)
            Tab1Widget.sname36.clicked.connect(self.n36fun)
            Tab1Widget.sname36.setGeometry(300,400,100,50)

            self.setStyleSheet('QPushButton{color:#5f5fff; background-color: #ffffff; padding: 0.3px 1px; margin: 2px 1px; font-weight: bold; border: solid; border-width: 1px; border-color: #5f5fff; border-radius: 3px}' 'QPushButton:hover {color: #fff; background-color: #5f5fff;}')


    def n1fun(self):
        global allplice
        global sp1tag
        global snametag1
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E27"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E27"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp1tag))
            Tab1Widget.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag1)
            one = str(sp1tag)
            display_list = [str(time), string, one]
            Tab1Widget.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K3"].value
            aimpoint = analize["J3"].value
            analize["C3"].value = snametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp1tag))
            analize["K3"].value = inputcode
            analize["J3"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag1)+":"+"￥"+str(sp1tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag1+'が残り少なくなっています')







        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n2fun(self):
        global allplice
        global sp2tag
        global snametag2
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E28"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E28"].value = zaiko1
            tag.save(setpath)


            allplice = (int(allplice)+int(sp2tag))
            Tab1Widget.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag2)
            one = str(sp2tag)
            display_list = [str(time), string, one]
            Tab1Widget.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K4"].value
            aimpoint = analize["J4"].value
            analize["I4"].value = snametag2
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp2tag))
            analize["K4"].value = inputcode
            analize["J4"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag2)+":"+"￥"+str(sp2tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag2+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n3fun(self):
        global allplice
        global sp3tag
        global snametag3
        global setpath
        global analizepath
        global info


        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E29"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E29"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp3tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag3)
            one = str(sp3tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K5"].value
            aimpoint = analize["J5"].value
            analize["I5"].value = snametag3
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp3tag))
            analize["K5"].value = inputcode
            analize["J5"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag3)+":"+"￥"+str(sp3tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag3+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n4fun(self):
        global allplice
        global sp4tag
        global snametag4
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E30"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E30"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp4tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag4)
            one = str(sp4tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K6"].value
            aimpoint = analize["J6"].value
            analize["I6"].value = snametag4
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp4tag))
            analize["K6"].value = inputcode
            analize["J6"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag4)+":"+"￥"+str(p4tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag41+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n5fun(self):
        global allplice
        global sp5tag
        global snametag5
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E31"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E31"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp5tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag5)
            one = str(sp5tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K7"].value
            aimpoint = analize["J7"].value
            analize["I7"].value = snametag5
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp5tag))
            analize["K7"].value = inputcode
            analize["J7"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag5)+":"+"￥"+str(p5tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag5+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n6fun(self):
        global allplice
        global sp6tag
        global snametag6
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E32"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E32"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp6tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag6)
            one = str(p6tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K8"].value
            aimpoint = analize["J8"].value
            analize["I8"].value = snametag6
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp6tag))
            analize["K8"].value = inputcode
            analize["J8"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag6)+":"+"￥"+str(sp6tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag6+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')

    def n7fun(self):
        global allplice
        global sp7tag
        global snametag7
        global setpath
        global info
        global analizepath

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E33"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E33"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp7tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag7)
            one = str(p7tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K9"].value
            aimpoint = analize["J9"].value
            analize["I9"].value = snametag7
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp7tag))
            analize["K9"].value = inputcode
            analize["J9"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag7)+":"+"￥"+str(sp7tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag7+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n8fun(self):
        global allplice
        global sp8tag
        global snametag8
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E34"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E34"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp8tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag8)
            one = str(p8tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K10"].value
            aimpoint = analize["J10"].value
            analize["I10"].value =snametag8
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp8tag))
            analize["K10"].value = inputcode
            analize["J10"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag8)+":"+"￥"+str(sp8tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag8+'が残り少なくなっています')


        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')

    def n9fun(self):
        global allplice
        global sp9tag
        global snametag9
        global info
        global analizepath
        global setpath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E35"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E35"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp9tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag9)
            one = str(p9tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K11"].value
            aimpoint = analize["J11"].value
            analize["I11"].value =snametag9
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp9tag))
            analize["K11"].value = inputcode
            analize["J11"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag9)+":"+"￥"+str(sp9tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag9+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n10fun(self):
        global allplice
        global sp10tag
        global snametag10
        global analizepath
        global setpath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E36"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E36"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp10tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag10)
            one = str(p10tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K12"].value
            aimpoint = analize["J12"].value
            analize["I12"].value =snametag10
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp10tag))
            analize["K12"].value = inputcode
            analize["J12"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag10)+":"+"￥"+str(sp10tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag10+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n11fun(self):
        global allplice
        global sp11tag
        global snametag11
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E37"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E37"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp11tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag11)
            one = str(p11tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K13"].value
            aimpoint = analize["J13"].value
            analize["I13"].value =snametag11
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp11tag))
            analize["K13"].value = inputcode
            analize["J13"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag11)+":"+"￥"+str(sp11tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag11+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n12fun(self):
        global allplice
        global info
        global sp12tag
        global snametag12
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E38"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E38"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp12tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag12)
            one = str(p12tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K14"].value
            aimpoint = analize["J14"].value
            analize["I14"].value =snametag12
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp12tag))
            analize["K14"].value = inputcode
            analize["J14"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag12)+":"+"￥"+str(sp12tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag12+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n13fun(self):
        global sp13tag
        global allplice
        global snametag13
        global info
        global analizepath
        global setpath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E39"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E39"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp13tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag13)
            one = str(sp13tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K15"].value
            aimpoint = analize["J15"].value
            analize["I15"].value = snametag13
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp13tag))
            analize["K15"].value = inputcode
            analize["J15"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag13)+":"+"￥"+str(sp13tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag13+'が残り少なくなっています')


        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')

    def n14fun(self):
        global allplice
        global sp14tag
        global snametag14
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E40"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E40"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp14tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag14)
            one = str(sp14tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K16"].value
            aimpoint = analize["J16"].value
            analize["I16"].value = snametag14
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp14tag))
            analize["K16"].value = inputcode
            analize["J16"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag14)+":"+"￥"+str(p14tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag14+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n15fun(self):
        global allplice
        global sp15tag
        global snametag15
        global analizepath
        global setpath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E41"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E41"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp15tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag15)
            one = str(sp15tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K17"].value
            aimpoint = analize["J17"].value
            analize["I17"].value = snametag15
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp15tag))
            analize["K17"].value = inputcode
            analize["J17"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag15)+":"+"￥"+str(p15tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag15+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n16fun(self):
        global allplice
        global sp16tag
        global snametag16
        global analizepath
        global setpath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E42"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E42"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp16tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag16)
            one = str(sp16tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K18"].value
            aimpoint = analize["J18"].value
            analize["I18"].value = snametag16
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp16tag))
            analize["K18"].value = inputcode
            analize["J18"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag16)+":"+"￥"+str(p16tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag16+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n17fun(self):
        global allplice
        global sp17tag
        global snametag17
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E43"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E43"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp17tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag17)
            one = str(sp17tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K19"].value
            aimpoint = analize["J19"].value
            analize["I19"].value = snametag17
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp17tag))
            analize["K19"].value = inputcode
            analize["J19"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag17)+":"+"￥"+str(p17tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag17+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n18fun(self):
        global allplice
        global sp18tag
        global snametag18
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E44"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E44"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp18tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag18)
            one = str(sp18tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K20"].value
            aimpoint = analize["J20"].value
            analize["I20"].value = snametag18
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp18tag))
            analize["K20"].value = inputcode
            analize["J20"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag18)+":"+"￥"+str(sp18tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag18+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n19fun(self):
        global allplice
        global sp19tag
        global snametag19
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k27"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k27"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp19tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag19)
            one = str(sp19tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K21"].value
            aimpoint = analize["J21"].value
            analize["I21"].value = snametag19
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp19tag))
            analize["K21"].value = inputcode
            analize["J21"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag19)+":"+"￥"+str(sp19tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag19+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n20fun(self):
        global allplice
        global sp20tag
        global snametag20
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k28"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k28"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp20tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag20)
            one = str(sp20tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K22"].value
            aimpoint = analize["J22"].value
            analize["I22"].value = snametag20
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp20tag))
            analize["K22"].value = inputcode
            analize["J22"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag20)+":"+"￥"+str(sp20tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag20+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n21fun(self):
        global allplice
        global sp21tag
        global snametag21
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k29"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k29"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp21tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag21)
            one = str(sp21tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K23"].value
            aimpoint = analize["J23"].value
            analize["I23"].value = snametag21
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp21tag))
            analize["K23"].value = inputcode
            analize["J23"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag21)+":"+"￥"+str(sp21tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag21+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n22fun(self):
        global allplice
        global sp22tag
        global snametag22
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k30"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k30"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(sp22tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag22)
            one = str(sp22tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K24"].value
            aimpoint = analize["J24"].value
            analize["I24"].value = nametag22
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp22tag))
            analize["K24"].value = inputcode
            analize["J24"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag22)+":"+"￥"+str(p22tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag22+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n23fun(self):
        global allplice
        global sp23tag
        global snametag23
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k31"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k31"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp23tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag23)
            one = str(sp23tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K25"].value
            aimpoint = analize["J25"].value
            analize["I25"].value = snametag23
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp23tag))
            analize["K25"].value = inputcode
            analize["J25"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag23)+":"+"￥"+str(sp23tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag23+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n24fun(self):
        global allplice
        global sp24tag
        global snametag24
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k32"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k32"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(sp24tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag24)
            one = str(sp24tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K26"].value
            aimpoint = analize["J26"].value
            analize["I3"].value = snametag24
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp24tag))
            analize["K26"].value = inputcode
            analize["J26"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag24)+":"+"￥"+str(sp24tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag24+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n25fun(self):
        global allplice
        global sp25tag
        global snametag25
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k33"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k33"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(sp25tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag25)
            one = str(sp25tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K27"].value
            aimpoint = analize["D27"].value
            analize["I27"].value = snametag25
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp25tag))
            analize["K27"].value = inputcode
            analize["D27"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag25)+":"+"￥"+str(sp25tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag25+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n26fun(self):
        global allplice
        global sp26tag
        global snametag26
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k34"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k34"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp26tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag26)
            one = str(sp26tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K28"].value
            aimpoint = analize["J28"].value
            analize["I28"].value = snametag26
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp26tag))
            analize["K28"].value = inputcode
            analize["J28"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag26)+":"+"￥"+str(sp26tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag26+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n27fun(self):
        global allplice
        global sp27tag
        global snametag27
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k35"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k35"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp27tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag27)
            one = str(sp27tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K29"].value
            aimpoint = analize["J29"].value
            analize["I29"].value = snametag27
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp27tag))
            analize["K29"].value = inputcode
            analize["J29"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag27)+":"+"￥"+str(sp27tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag27+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n28fun(self):
        global allplice
        global sp28tag
        global snametag28
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k36"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k36"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp28tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag28)
            one = str(sp28tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K30"].value
            aimpoint = analize["J30"].value
            analize["I30"].value = snametag28
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp28tag))
            analize["K30"].value = inputcode
            analize["J30"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag28)+":"+"￥"+str(sp28tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag28+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n29fun(self):
        global allplice
        global sp29tag
        global snametag29
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k37"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k37"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(sp29tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag29)
            one = str(sp29tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K31"].value
            aimpoint = analize["J31"].value
            analize["I31"].value = snametag29
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp29tag))
            analize["K31"].value = inputcode
            analize["J31"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag29)+":"+"￥"+str(sp29tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag29+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n30fun(self):
        global allplice
        global sp30tag
        global snametag30
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k38"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k38"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(sp30tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag30)
            one = str(sp30tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K32"].value
            aimpoint = analize["J32"].value
            analize["I32"].value = snametag30
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp30tag))
            analize["K32"].value = inputcode
            analize["J32"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag30)+":"+"￥"+str(sp30tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag30+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n31fun(self):
        global allplice
        global analizepath
        global sp31tag
        global info
        global snametag31
        global setpath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k39"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k39"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(sp31tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag31)
            one = str(sp31tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K33"].value
            aimpoint = analize["J33"].value
            analize["I33"].value = snametag31
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp31tag))
            analize["K33"].value = inputcode
            analize["J33"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag31)+":"+"￥"+str(sp31tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag31+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n32fun(self):
        global allplice
        global sp32tag
        global info
        global snametag32
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k40"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k40"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(sp32tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag32)
            one = str(sp32tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K34"].value
            aimpoint = analize["J34"].value
            analize["I34"].value = snametag32
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp32tag))
            analize["K34"].value = inputcode
            analize["J34"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag32)+":"+"￥"+str(sp32tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag32+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n33fun(self):
        global allplice
        global sp33tag
        global snametag33
        global info
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k41"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k41"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p33tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag33)
            one = str(p33tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K35"].value
            aimpoint = analize["J35"].value
            analize["I35"].value = nametag33
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p33tag))
            analize["K35"].value = inputcode
            analize["J35"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag33)+":"+"￥"+str(p33tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag33+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n34fun(self):
        global allplice
        global sp34tag
        global snametag34
        global info
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k42"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k42"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(sp34tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag34)
            one = str(sp34tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            ansatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K36"].value
            aimpoint = analize["J36"].value
            analize["I36"].value = snametag34
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp34tag))
            analize["K36"].value = inputcode
            analize["J36"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag34)+":"+"￥"+str(sp34tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag34+'が残り少なくなっています')


        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n35fun(self):
        global allplice
        global sp35tag
        global snametag35
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k43"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k43"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(sp35tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag35)
            one = str(sp35tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K37"].value
            aimpoint = analize["J37"].value
            analize["I37"].value = snametag35
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp35tag))
            analize["K37"].value = inputcode
            analize["J37"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag35)+":"+"￥"+str(sp35tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag35+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n36fun(self):
        global allplice
        global sp36tag
        global info
        global snametag36
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k44"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k44"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(sp36tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(snametag36)
            one = str(sp36tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["K38"].value
            aimpoint = analize["J38"].value
            analize["I38"].value = snametag36
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(sp36tag))
            analize["K38"].value = inputcode
            analize["J38"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(snametag36)+":"+"￥"+str(sp36tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', snametag36+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


class Tab3Widget(QWidget):
    def __init__(self, parent=None):
        super(Tab3Widget, self).__init__(parent)

        Tab1Widget.dname1 = QPushButton('indname1', self)#//Button
        Tab1Widget.dname1.clicked.connect(self.n1fun)
        Tab1Widget.dname1.setGeometry(0,0,100,50)
        Tab1Widget.dname2 = QPushButton('indname2', self)
        Tab1Widget.dname2.clicked.connect(self.n2fun)
        Tab1Widget.dname2.setGeometry(0,50,100,50)
        Tab1Widget.dname3 = QPushButton('indname3', self)
        Tab1Widget.dname3.clicked.connect(self.n3fun)
        Tab1Widget.dname3.setGeometry(0,100,100,50)
        Tab1Widget.dname4 = QPushButton('indname4', self)
        Tab1Widget.dname4.clicked.connect(self.n4fun)
        Tab1Widget.dname4.setGeometry(0,150,100,50)
        Tab1Widget.dname5 = QPushButton('indnam5', self)
        Tab1Widget.dname5.clicked.connect(self.n5fun)
        Tab1Widget.dname5.setGeometry(0,200,100,50)
        Tab1Widget.dname6 = QPushButton('indname6', self)
        Tab1Widget.dname6.clicked.connect(self.n6fun)
        Tab1Widget.dname6.setGeometry(0,250,100,50)
        Tab1Widget.dname7 = QPushButton('indname7', self)
        Tab1Widget.dname7.clicked.connect(self.n7fun)
        Tab1Widget.dname7.setGeometry(0,300,100,50)
        Tab1Widget.dname8 = QPushButton('indname8', self)
        Tab1Widget.dname8.clicked.connect(self.n8fun)
        Tab1Widget.dname8.setGeometry(0,350,100,50)
        Tab1Widget.dname9 = QPushButton('indname9', self)
        Tab1Widget.dname9.clicked.connect(self.n9fun)
        Tab1Widget.dname9.setGeometry(0,400,100,50)

        Tab1Widget.dname10 = QPushButton('indname10', self)
        Tab1Widget.dname10.clicked.connect(self.n10fun)
        Tab1Widget.dname10.setGeometry(100,0,100,50)
        Tab1Widget.dname11 = QPushButton('indname11', self)
        Tab1Widget.dname11.clicked.connect(self.n11fun)
        Tab1Widget.dname11.setGeometry(100,50,100,50)
        Tab1Widget.dname12 = QPushButton('indname12', self)
        Tab1Widget.dname12.clicked.connect(self.n12fun)
        Tab1Widget.dname12.setGeometry(100,100,100,50)
        Tab1Widget.dname13 = QPushButton('indname13', self)
        Tab1Widget.dname13.clicked.connect(self.n13fun)
        Tab1Widget.dname13.setGeometry(100,150,100,50)
        Tab1Widget.dname14 = QPushButton('indnam14', self)
        Tab1Widget.dname14.clicked.connect(self.n14fun)
        Tab1Widget.dname14.setGeometry(100,200,100,50)
        Tab1Widget.dname15 = QPushButton('indname15', self)
        Tab1Widget.dname15.clicked.connect(self.n15fun)
        Tab1Widget.dname15.setGeometry(100,250,100,50)
        Tab1Widget.dname16= QPushButton('indname16', self)
        Tab1Widget.dname16.clicked.connect(self.n16fun)
        Tab1Widget.dname16.setGeometry(100,300,100,50)
        Tab1Widget.dname17 = QPushButton('indname17', self)
        Tab1Widget.dname17.clicked.connect(self.n17fun)
        Tab1Widget.dname17.setGeometry(100,350,100,50)
        Tab1Widget.dname18 = QPushButton('indname18', self)
        Tab1Widget.dname18.clicked.connect(self.n18fun)
        Tab1Widget.dname18.setGeometry(100,400,100,50)

        Tab1Widget.dname19 = QPushButton('indname1', self)#//Button
        Tab1Widget.dname19.clicked.connect(self.n19fun)
        Tab1Widget.dname19.setGeometry(200,0,100,50)
        Tab1Widget.dname20 = QPushButton('indname2', self)
        Tab1Widget.dname20.clicked.connect(self.n20fun)
        Tab1Widget.dname20.setGeometry(200,50,100,50)
        Tab1Widget.dname21 = QPushButton('indname3', self)
        Tab1Widget.dname21.clicked.connect(self.n21fun)
        Tab1Widget.dname21.setGeometry(200,100,100,50)
        Tab1Widget.dname22 = QPushButton('indname4', self)
        Tab1Widget.dname22.clicked.connect(self.n22fun)
        Tab1Widget.dname22.setGeometry(200,150,100,50)
        Tab1Widget.dname23 = QPushButton('indnam5', self)
        Tab1Widget.dname23.clicked.connect(self.n23fun)
        Tab1Widget.dname23.setGeometry(200,200,100,50)
        Tab1Widget.dname24 = QPushButton('indname6', self)
        Tab1Widget.dname24.clicked.connect(self.n24fun)
        Tab1Widget.dname24.setGeometry(200,250,100,50)
        Tab1Widget.dname25 = QPushButton('indname7', self)
        Tab1Widget.dname25.clicked.connect(self.n25fun)
        Tab1Widget.dname25.setGeometry(200,300,100,50)
        Tab1Widget.dname26 = QPushButton('indname8', self)
        Tab1Widget.dname26.clicked.connect(self.n26fun)
        Tab1Widget.dname26.setGeometry(200,350,100,50)
        Tab1Widget.dname27 = QPushButton('indname9', self)
        Tab1Widget.dname27.clicked.connect(self.n27fun)
        Tab1Widget.dname27.setGeometry(200,400,100,50)

        Tab1Widget.dname28 = QPushButton('indname10', self)
        Tab1Widget.dname28.clicked.connect(self.n28fun)
        Tab1Widget.dname28.setGeometry(300,0,100,50)
        Tab1Widget.dname29 = QPushButton('indname11', self)
        Tab1Widget.dname29.clicked.connect(self.n29fun)
        Tab1Widget.dname29.setGeometry(300,50,100,50)
        Tab1Widget.dname30 = QPushButton('indname12', self)
        Tab1Widget.dname30.clicked.connect(self.n30fun)
        Tab1Widget.dname30.setGeometry(300,100,100,50)
        Tab1Widget.dname31 = QPushButton('indname13', self)
        Tab1Widget.dname31.clicked.connect(self.n31fun)
        Tab1Widget.dname31.setGeometry(300,150,100,50)
        Tab1Widget.dname32 = QPushButton('indnam14', self)
        Tab1Widget.dname32.clicked.connect(self.n32fun)
        Tab1Widget.dname32.setGeometry(300,200,100,50)
        Tab1Widget.dname33 = QPushButton('indname15', self)
        Tab1Widget.dname33.clicked.connect(self.n33fun)
        Tab1Widget.dname33.setGeometry(300,250,100,50)
        Tab1Widget.dname34= QPushButton('indname16', self)
        Tab1Widget.dname34.clicked.connect(self.n34fun)
        Tab1Widget.dname34.setGeometry(300,300,100,50)
        Tab1Widget.dname35 = QPushButton('indname17', self)
        Tab1Widget.dname35.clicked.connect(self.n35fun)
        Tab1Widget.dname35.setGeometry(300,350,100,50)
        Tab1Widget.dname36 = QPushButton('indname18', self)
        Tab1Widget.dname36.clicked.connect(self.n36fun)
        Tab1Widget.dname36.setGeometry(300,400,100,50)

        self.setStyleSheet('QPushButton{color:#5f5fff; background-color: #ffffff; padding: 0.3px 1px; margin: 2px 1px; font-weight: bold; border: solid; border-width: 1px; border-color: #5f5fff; border-radius: 3px}' 'QPushButton:hover {color: #fff; background-color: #5f5fff;}')



    def n1fun(self):
        global allplice
        global dp1tag
        global dnametag1
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E50"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E50"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp1tag))
            Tab1Widget.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag1)
            one = str(dp1tag)
            display_list = [str(time), string, one]
            Tab1Widget.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q3"].value
            aimpoint = analize["P3"].value
            analize["O3"].value = dnametag1
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp1tag))
            analize["Q3"].value = inputcode
            analize["P3"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag1)+":"+"￥"+str(dp1tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag1+'が残り少なくなっています')







        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n2fun(self):
        global allplice
        global dp2tag
        global dnametag2
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E51"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E51"].value = zaiko1
            tag.save(setpath)


            allplice = (int(allplice)+int(dp2tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag2)
            one = str(dp2tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q4"].value
            aimpoint = analize["P4"].value
            analize["O4"].value = dnametag2
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp2tag))
            analize["Q4"].value = inputcode
            analize["P4"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag2)+":"+"￥"+str(dp2tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag2+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n3fun(self):
        global allplice
        global dp3tag
        global dnametag3
        global setpath
        global analizepath
        global info


        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E52"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E52"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp3tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag3)
            one = str(dp3tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q5"].value
            aimpoint = analize["P5"].value
            analize["O5"].value = dnametag3
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp3tag))
            analize["Q5"].value = inputcode
            analize["P5"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag3)+":"+"￥"+str(dp3tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag3+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n4fun(self):
        global allplice
        global dp4tag
        global dnametag4
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E53"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E53"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp4tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag4)
            one = str(dp4tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q6"].value
            aimpoint = analize["P6"].value
            analize["O6"].value = dnametag4
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp4tag))
            analize["Q6"].value = inputcode
            analize["P6"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag4)+":"+"￥"+str(p4tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag41+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n5fun(self):
        global allplice
        global dp5tag
        global dnametag5
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E54"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E54"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp5tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag5)
            one = str(dp5tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q7"].value
            aimpoint = analize["P7"].value
            analize["O7"].value = dnametag5
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp5tag))
            analize["Q7"].value = inputcode
            analize["P7"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag5)+":"+"￥"+str(p5tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag5+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n6fun(self):
        global allplice
        global dp6tag
        global dnametag6
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E55"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E55"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp6tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag6)
            one = str(dp6tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q8"].value
            aimpoint = analize["P8"].value
            analize["O8"].value = dnametag6
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp6tag))
            analize["Q8"].value = inputcode
            analize["P8"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag6)+":"+"￥"+str(dp6tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag6+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')

    def n7fun(self):
        global allplice
        global dp7tag
        global dnametag7
        global setpath
        global info
        global analizepath

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E56"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E56"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(p7tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(nametag7)
            one = str(p7tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q9"].value
            aimpoint = analize["P9"].value
            analize["O9"].value = nametag7
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p7tag))
            analize["Q9"].value = inputcode
            analize["P9"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(nametag7)+":"+"￥"+str(p7tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag7+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n8fun(self):
        global allplice
        global dp8tag
        global dnametag8
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E57"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E57"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp8tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag8)
            one = str(dp8tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q10"].value
            aimpoint = analize["P10"].value
            analize["O10"].value = dnametag8
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp8tag))
            analize["Q10"].value = inputcode
            analize["P10"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag8)+":"+"￥"+str(dp8tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag8+'が残り少なくなっています')


        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')

    def n9fun(self):
        global allplice
        global dp9tag
        global dnametag9
        global info
        global analizepath
        global setpath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E58"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E58"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp9tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag9)
            one = str(dp9tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q11"].value
            aimpoint = analize["P11"].value
            analize["O11"].value = dnametag9
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp9tag))
            analize["Q11"].value = inputcode
            analize["P11"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag9)+":"+"￥"+str(dp9tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag9+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n10fun(self):
        global allplice
        global dp10tag
        global dnametag10
        global analizepath
        global setpath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E59"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E59"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp10tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag10)
            one = str(dp10tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q12"].value
            aimpoint = analize["P12"].value
            analize["O12"].value = dnametag10
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp10tag))
            analize["Q12"].value = inputcode
            analize["P12"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag10)+":"+"￥"+str(dp10tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag10+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n11fun(self):
        global allplice
        global dp11tag
        global dnametag11
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E60"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E60"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp11tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag11)
            one = str(dp11tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q13"].value
            aimpoint = analize["P13"].value
            analize["O13"].value = dnametag11
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp11tag))
            analize["Q13"].value = inputcode
            analize["P13"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag11)+":"+"￥"+str(dp11tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag11+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n12fun(self):
        global allplice
        global info
        global dp12tag
        global dnametag12
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E61"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E61"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp12tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag12)
            one = str(dp12tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q14"].value
            aimpoint = analize["P14"].value
            analize["O14"].value = dnametag12
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp12tag))
            analize["Q14"].value = inputcode
            analize["P14"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag12)+":"+"￥"+str(dp12tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag12+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n13fun(self):
        global dp13tag
        global dallplice
        global dnametag13
        global info
        global analizepath
        global setpath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E62"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E62"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp13tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag13)
            one = str(dp13tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q15"].value
            aimpoint = analize["P15"].value
            analize["O15"].value = dnametag13
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp13tag))
            analize["Q15"].value = inputcode
            analize["P15"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag13)+":"+"￥"+str(dp13tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag13+'が残り少なくなっています')


        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')

    def n14fun(self):
        global allplice
        global dp14tag
        global dnametag14
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E63"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E63"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp14tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag14)
            one = str(dp14tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q16"].value
            aimpoint = analize["P16"].value
            analize["O16"].value = dnametag14
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp14tag))
            analize["Q16"].value = inputcode
            analize["P16"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag14)+":"+"￥"+str(dp14tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag14+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n15fun(self):
        global allplice
        global dp15tag
        global dnametag15
        global analizepath
        global setpath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E64"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E64"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp15tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag15)
            one = str(dp15tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q17"].value
            aimpoint = analize["P17"].value
            analize["O17"].value = dnametag15
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp15tag))
            analize["Q17"].value = inputcode
            analize["P17"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag15)+":"+"￥"+str(p15tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag15+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n16fun(self):
        global allplice
        global dp16tag
        global dnametag16
        global analizepath
        global setpath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E65"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E65"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp16tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag16)
            one = str(dp16tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q18"].value
            aimpoint = analize["P18"].value
            analize["O18"].value = dnametag16
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp16tag))
            analize["Q18"].value = inputcode
            analize["P18"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag16)+":"+"￥"+str(p16tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag16+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n17fun(self):
        global allplice
        global dp17tag
        global dnametag17
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E66"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E66"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp17tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag17)
            one = str(dp17tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q19"].value
            aimpoint = analize["P19"].value
            analize["O19"].value = dnametag17
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp17tag))
            analize["Q19"].value = inputcode
            analize["P19"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag17)+":"+"￥"+str(p17tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag17+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n18fun(self):
        global allplice
        global dp18tag
        global dnametag18
        global setpath
        global analizepath
        global info

        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["E67"].value
            zaiko1 = int(zaiko1) - int(1)
            act["E67"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp18tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag18)
            one = str(dp18tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q20"].value
            aimpoint = analize["P20"].value
            analize["O20"].value = dnametag18
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp18tag))
            analize["Q20"].value = inputcode
            analize["P20"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag18)+":"+"￥"+str(dp18tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag18+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n19fun(self):
        global allplice
        global dp19tag
        global dnametag19
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k50"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k50"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp19tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag19)
            one = strd(p19tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q21"].value
            aimpoint = analize["P21"].value
            analize["O21"].value = dnametag19
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(p19tag))
            analize["Q21"].value = inputcode
            analize["P21"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag19)+":"+"￥"+str(dp19tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag19+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')


    def n20fun(self):
        global allplice
        global dp20tag
        global dnametag20
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k51"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k51"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp20tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag20)
            one = str(dp20tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q22"].value
            aimpoint = analize["P22"].value
            analize["O22"].value = dnametag20
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp20tag))
            analize["Q22"].value = inputcode
            analize["P22"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag20)+":"+"￥"+str(dp20tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag20+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n21fun(self):
        global allplice
        global dp21tag
        global dnametag21
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k52"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k52"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp21tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag21)
            one = str(dp21tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q23"].value
            aimpoint = analize["D23"].value
            analize["O23"].value = dnametag21
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp21tag))
            analize["Q23"].value = inputcode
            analize["D23"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag21)+":"+"￥"+str(dp21tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag21+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n22fun(self):
        global allplice
        global dp22tag
        global dnametag22
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k53"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k53"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(dp22tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag22)
            one = str(dp22tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q24"].value
            aimpoint = analize["P24"].value
            analize["O24"].value = dnametag22
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp22tag))
            analize["Q24"].value = inputcode
            analize["P24"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag22)+":"+"￥"+str(dp22tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag22+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n23fun(self):
        global allplice
        global dp23tag
        global dnametag23
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k54"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k54"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp23tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag23)
            one = str(dp23tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q25"].value
            aimpoint = analize["P25"].value
            analize["O25"].value = dnametag23
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp23tag))
            analize["Q25"].value = inputcode
            analize["P25"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag23)+":"+"￥"+str(dp23tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', nametag23+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n24fun(self):
        global allplice
        global dp24tag
        global dnametag24
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k55"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k55"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(dp24tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag24)
            one = str(dp24tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q26"].value
            aimpoint = analize["P26"].value
            analize["O26"].value = dnametag24
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp24tag))
            analize["Q26"].value = inputcode
            analize["P26"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag24)+":"+"￥"+str(dp24tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag24+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n25fun(self):
        global allplice
        global dp25tag
        global dnametag25
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k56"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k56"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(dp25tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag25)
            one = str(dp25tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q27"].value
            aimpoint = analize["P27"].value
            analize["O27"].value = dnametag25
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp25tag))
            analize["Q27"].value = inputcode
            analize["P27"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag25)+":"+"￥"+str(dp25tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag25+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n26fun(self):
        global allplice
        global dp26tag
        global dnametag26
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k57"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k57"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp26tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag26)
            one = str(dp26tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q28"].value
            aimpoint = analize["P28"].value
            analize["O28"].value = dnametag26
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp26tag))
            analize["Q28"].value = inputcode
            analize["P28"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag26)+":"+"￥"+str(dp26tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag26+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n27fun(self):
        global allplice
        global dp27tag
        global dnametag27
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k58"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k58"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp27tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag27)
            one = str(dp27tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q29"].value
            aimpoint = analize["P29"].value
            analize["O29"].value = dnametag27
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp27tag))
            analize["Q29"].value = inputcode
            analize["P29"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag27)+":"+"￥"+str(dp27tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag27+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n28fun(self):
        global allplice
        global dp28tag
        global dnametag28
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k59"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k59"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp28tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag28)
            one = str(dp28tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q30"].value
            aimpoint = analize["P30"].value
            analize["O30"].value = dnametag28
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp28tag))
            analize["Q30"].value = inputcode
            analize["P30"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag28)+":"+"￥"+str(dp28tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag28+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n29fun(self):
        global allplice
        global dp29tag
        global dnametag29
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k60"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k60"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(dp29tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag29)
            one = str(dp29tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q31"].value
            aimpoint = analize["P31"].value
            analize["O31"].value = dnametag29
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp29tag))
            analize["Q31"].value = inputcode
            analize["P31"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag29)+":"+"￥"+str(dp29tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag29+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n30fun(self):
        global allplice
        global dp30tag
        global dnametag30
        global setpath
        global info
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k61"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k61"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(dp30tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag30)
            one = str(dp30tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q32"].value
            aimpoint = analize["P32"].value
            analize["O32"].value = dnametag30
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp30tag))
            analize["Q32"].value = inputcode
            analize["P32"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag30)+":"+"￥"+str(dp30tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag30+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')





    def n31fun(self):
        global allplice
        global analizepath
        global dp31tag
        global info
        global dnametag31
        global setpath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k62"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k62"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(dp31tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag31)
            one = str(dp31tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q33"].value
            aimpoint = analize["P33"].value
            analize["O33"].value = dnametag31
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp31tag))
            analize["Q33"].value = inputcode
            analize["P33"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag31)+":"+"￥"+str(dp31tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag31+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n32fun(self):
        global allplice
        global dp32tag
        global info
        global dnametag32
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k63"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k63"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(dp32tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag32)
            one = str(dp32tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q34"].value
            aimpoint = analize["P34"].value
            analize["O34"].value = dnametag32
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp32tag))
            analize["Q34"].value = inputcode
            analize["P34"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag32)+":"+"￥"+str(pd32tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag32+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n33fun(self):
        global allplice
        global dp33tag
        global dnametag33
        global info
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k64"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k64"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp33tag))
            self.plicelabel.setText(str(allpliced))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag33)
            one = str(dp33tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q35"].value
            aimpoint = analize["P35"].value
            analize["O35"].value = dnametag33
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp33tag))
            analize["Q35"].value = inputcode
            analize["P35"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag33)+":"+"￥"+str(dp33tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag33+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n34fun(self):
        global allplice
        global dp34tag
        global dnametag34
        global info
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k65"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k65"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(dp34tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag34)
            one = str(dp34tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q36"].value
            aimpoint = analize["P36"].value
            analize["O36"].value = dnametag34
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp34tag))
            analize["Q36"].value = inputcode
            analize["P36"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag34)+":"+"￥"+str(dp34tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag34+'が残り少なくなっています')


        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')



    def n35fun(self):
        global allplice
        global dp35tag
        global dnametag35
        global setpath
        global analizepath
        global info
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k66"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k66"].value = zaiko1
            tag.save(setpath)
            allplice = (int(allplice)+int(dp35tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag35)
            one = str(dp35tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q37"].value
            aimpoint = analize["P37"].value
            analize["O37"].value = dnametag35
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp35tag))
            analize["Q37"].value = inputcode
            analize["P37"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag35)+":"+"￥"+str(dp35tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()
            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag35+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')




    def n36fun(self):
        global allplice
        global dp36tag
        global info
        global dnametag36
        global setpath
        global analizepath
        try:
            tag = openpyxl.load_workbook(setpath)
            act = tag.active
            zaiko1 = act["k67"].value
            zaiko1 = int(zaiko1) - int(1)
            act["k67"].value = zaiko1
            tag.save(setpath)

            allplice = (int(allplice)+int(dp36tag))
            self.plicelabel.setText(str(allplice))
            t = datetime.datetime.now()
            time = t.time()
            string = str(dnametag36)
            one = str(dp36tag)
            display_list = [str(time), string, one]
            self.model.additeminrow(display_list)

            anatag = openpyxl.load_workbook(analizepath)
            analize = anatag.active
            inputcode = analize["Q38"].value
            aimpoint = analize["P38"].value
            analize["O38"].value = dnametag36
            point = (int(aimpoint)+int(1))
            inputcode = (int(inputcode)+int(dp36tag))
            analize["Q38"].value = inputcode
            analize["P38"].value = point
            anatag.save(analizepath)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":"+str(dnametag36)+":"+"￥"+str(dp36tag)+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()

            if int(zaiko1) <= int(info):
                massage = QMessageBox.about(self, '確認', dnametag36+'が残り少なくなっています')

        except TypeError:
            dialog = QMessageBox.warning(self, 'エラー', '商品情報を入力してください')













class MainWin(QMainWindow):
    def __init__(self, parent=None):
        super(MainWin, self).__init__(parent)#GUI


        Tab1Widget.plicelabel = QLabel(self)
        Tab1Widget.plicelabel.setFrameStyle(QFrame.Box)
        Tab1Widget.plicelabel.setStyleSheet("font-size:50px; background-color:black; color:white;")
        Tab1Widget.plicelabel.setAlignment(Qt.AlignRight)
        Tab1Widget.plicelabel.setGeometry(800, 0, 500, 70) #label
        Tab1Widget.plicelabel.show()

        namelabel = QLabel('合計金額', self)
        namelabel.setStyleSheet("font-size:50px; font-family:fantasy;")
        namelabel.setGeometry(600, 0, 200, 70)

        Tab1Widget.paylabel = QLabel(self)
        Tab1Widget.paylabel.setFrameStyle(QFrame.Box)
        Tab1Widget.paylabel.setStyleSheet("font-size:25px; background-color:black; color:white;")
        Tab1Widget.paylabel.setAlignment(Qt.AlignRight)
        Tab1Widget.paylabel.setGeometry(800, 80, 200, 35)

        name2label = QLabel('お預金', self)
        name2label.setStyleSheet("font-size:25px; font-family:fantasy;")
        name2label.setGeometry(700, 80, 200, 35)

        Tab1Widget.changelabel = QLabel(self)
        Tab1Widget.changelabel.setFrameStyle(QFrame.Box)
        Tab1Widget.changelabel.setStyleSheet("font-size:25px; background-color:black; color:white;")
        Tab1Widget.changelabel.setAlignment(Qt.AlignRight)
        Tab1Widget.changelabel.setGeometry(800, 120, 200, 35)

        name3label = QLabel('お釣り', self)
        name3label.setStyleSheet("font-size:25px; font-family:fantasy;")
        name3label.setGeometry(700, 120, 200, 35)

        Tab1Widget.pay2label = QLabel(self)
        Tab1Widget.pay2label.setFrameStyle(QFrame.Box)
        Tab1Widget.pay2label.setStyleSheet("font-size:40px;")
        Tab1Widget.pay2label.setAlignment(Qt.AlignRight)
        Tab1Widget.pay2label.setGeometry(1045, 200, 250, 50)


        Tab1Widget.Gbutton = QPushButton('Setting', self)
        Tab1Widget.Gbutton.clicked.connect(self.settingfun)
        Tab1Widget.Gbutton.setGeometry(0, 100, 100, 50)

        Tab1Widget.kzeibutton = QPushButton('消費税(8%)', self)
        Tab1Widget.kzeibutton.clicked.connect(self.kzeifun)
        Tab1Widget.kzeibutton.setGeometry(0, 150, 100, 50)

        Tab1Widget.nzeibutton = QPushButton('消費税(10%)', self)
        Tab1Widget.nzeibutton.clicked.connect(self.nzeifun)
        Tab1Widget.nzeibutton.setGeometry(0, 200, 100, 50)

        Tab1Widget.nebutton = QPushButton('値引き', self)
        Tab1Widget.nebutton.clicked.connect(self.nebfun)
        Tab1Widget.nebutton.setGeometry(0, 250, 100, 50)



















        clear1 = QPushButton('clear', self)
        clear1.clicked.connect(self.clear1fun)
        clear1.setGeometry(910, 225, 100, 50)
        paycheckbutton = QPushButton('支払い', self)
        paycheckbutton.clicked.connect(self.payfun)
        paycheckbutton.setGeometry(910, 175, 100, 50)


        pay9 = QPushButton('9', self)
        pay9.clicked.connect(self.p9fun)
        pay9.setGeometry(1210, 250, 70, 70)
        pay8 = QPushButton('8', self)
        pay8.clicked.connect(self.p8fun)
        pay8.setGeometry(1140, 250, 70, 70)
        pay7 = QPushButton('7', self)
        pay7.clicked.connect(self.p7fun)
        pay7.setGeometry(1070, 250, 70, 70)
        pay6 = QPushButton('6', self)
        pay6.clicked.connect(self.p6fun)
        pay6.setGeometry(1210, 320, 70, 70)
        pay5 = QPushButton('5', self)
        pay5.clicked.connect(self.p5fun)
        pay5.setGeometry(1140, 320, 70, 70)
        pay4 = QPushButton('4', self)
        pay4.clicked.connect(self.p4fun)
        pay4.setGeometry(1070, 320, 70, 70)
        pay3 = QPushButton('3', self)
        pay3.clicked.connect(self.p3fun)
        pay3.setGeometry(1210, 390, 70, 70)
        pay2 = QPushButton('2', self)
        pay2.clicked.connect(self.p2fun)
        pay2.setGeometry(1140, 390, 70, 70)
        pay1 = QPushButton('1', self)
        pay1.clicked.connect(self.p1fun)
        pay1.setGeometry(1070, 390, 70, 70)
        pay00 = QPushButton('00', self)
        pay00.clicked.connect(self.p00fun)
        pay00.setGeometry(1210, 460, 70, 70)
        pay0 = QPushButton('0', self)
        pay0.clicked.connect(self.p0fun)
        pay0.setGeometry(1140, 460, 70, 70)
        pay000 = QPushButton('000', self)
        pay000.clicked.connect(self.p000fun)
        pay000.setGeometry(1070, 460, 70, 70)
        clear2 = QPushButton('clear', self)
        clear2.clicked.connect(self.clear2fun)
        clear2.setGeometry(1070, 530, 210, 70)




        Tab1Widget.model = Model(0, 3, self)
        self.mainlist = QTreeView(self)
        self.mainlist.setModel(Tab1Widget.model)
        self.mainlist.setColumnWidth(1, 140)
        self.mainlist.setGeometry(100, 100, 350, 550)


        self.tab = QTabWidget(self)
        self.tab.addTab(Tab1Widget(parent=None), 'メイン')
        self.tab.addTab(Tab2Widget(parent=None), 'サブ')
        self.tab.addTab(Tab3Widget(parent=None), 'サブ２')
        self.tab.setGeometry(500, 150, 405, 500)

















        p = self.palette()
        p.setColor(self.backgroundRole(), Qt.white)
        self.setPalette(p)
        self.setGeometry(300, 300, 600, 350)
        self.setWindowTitle('PyPOS')



#system


    def kzeifun(self):
        global allplice
        global zcode
        kzei = (int(allplice)*0.08)
        allplice = (int(allplice)+int(kzei))
        Tab1Widget.plicelabel.setText(str(allplice))
        zcode = ("8%")

    def nzeifun(self):
        global allplice
        nzei = (int(allplice)*0.1)
        allplice = (int(allplice)+int(nzei))
        Tab1Widget.plicelabel.setText(str(allplice))
        zcode = ("10%")

    def nebfun(self):
        global allplice
        text, indialog = QInputDialog.getText(self, '値引き', '値引き額を入力')
        if indialog:
            allplice = (int(allplice)-int(text))
            Tab1Widget.plicelabel.setText(str(allplice))



    def payfun(self):
        global paymanay
        global allplice
        global change
        global zcode
        change = (int(paymanay)-int(allplice))
        Tab1Widget.changelabel.setText(str(change))
        t = datetime.datetime.now()
        time = t.time()
        string = str('=====総額=======')
        one = str(allplice)
        display_list = [str(time), string, one]
        Tab1Widget.model.additeminrow(display_list)

        datatime = datetime.datetime.now()
        datasorce = (str(datatime)+":===総額===:"+"￥"+str(allplice)+"("+zcode+")"+"\n")
        data = os.path.abspath("取引データ.txt")
        file = open(data, 'a')
        file.write(datasorce)
        file.close()



#paysystem

    def p9fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('9'))
        Tab1Widget.paylabel.setText(str(paymanay))
        Tab1Widget.pay2label.setText(str(paymanay))

    def p8fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('8'))
        Tab1Widget.paylabel.setText(str(paymanay))
        Tab1Widget.pay2label.setText(str(paymanay))

    def p7fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('7'))
        Tab1Widget.paylabel.setText(str(paymanay))
        Tab1Widget.pay2label.setText(str(paymanay))

    def p6fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('6'))
        Tab1Widget.paylabel.setText(str(paymanay))
        Tab1Widget.pay2label.setText(str(paymanay))

    def p5fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('5'))
        Tab1Widget.paylabel.setText(str(paymanay))
        Tab1Widget.pay2label.setText(str(paymanay))

    def p4fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('4'))
        Tab1Widget.paylabel.setText(str(paymanay))
        Tab1Widget.pay2label.setText(str(paymanay))

    def p3fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('3'))
        Tab1Widget.paylabel.setText(str(paymanay))
        Tab1Widget.pay2label.setText(str(paymanay))

    def p2fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('2'))
        Tab1Widget.paylabel.setText(str(paymanay))
        Tab1Widget.pay2label.setText(str(paymanay))

    def p1fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('1'))
        Tab1Widget.paylabel.setText(str(paymanay))
        Tab1Widget.pay2label.setText(str(paymanay))

    def p00fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('00'))
        Tab1Widget.paylabel.setText(str(paymanay))
        Tab1Widget.pay2label.setText(str(paymanay))

    def p0fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('0'))
        Tab1Widget.paylabel.setText(str(paymanay))
        Tab1Widget.pay2label.setText(str(paymanay))

    def p000fun(self):
        global paymanay
        paymanay = (str(paymanay)+str('000'))
        Tab1Widget.paylabel.setText(str(paymanay))
        Tab1Widget.pay2label.setText(str(paymanay))

    def clear2fun(self):
        global paymanay
        paymanay = (str(''))
        Tab1Widget.paylabel.setText(str(paymanay))
        Tab1Widget.pay2label.setText(str(paymanay))
        t = datetime.datetime.now()

    def clear1fun(self):
        global allplice
        allplice = (int(0))
        Tab1Widget.plicelabel.setText(str(allplice))
        t = datetime.datetime.now()
        time = t.time()
        string = str("クリア")
        one = str("")
        display_list = [str(time), string, one]
        Tab1Widget.model.additeminrow(display_list)

        datatime = datetime.datetime.now()
        datasorce = (str(datatime)+":以下クリア:"+"￥0"+"\n")
        data = os.path.abspath("取引データ.txt")
        file = open(data, 'a')
        file.write(datasorce)
        file.close()

#Key
    def keyPressEvent(self, e):
        global paymanay
        global allplice
        global change

        if e.key() == Qt.Key_1:
            paymanay = (str(paymanay)+str('1'))
            Tab1Widget.paylabel.setText(str(paymanay))
            Tab1Widget.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_2:
            paymanay = (str(paymanay)+str('2'))
            Tab1Widget.paylabel.setText(str(paymanay))
            Tab1Widget.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_3:
            paymanay = (str(paymanay)+str('3'))
            Tab1Widget.paylabel.setText(str(paymanay))
            Tab1Widget.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_4:
            paymanay = (str(paymanay)+str('4'))
            Tab1Widget.paylabel.setText(str(paymanay))
            Tab1Widget.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_5:
            paymanay = (str(paymanay)+str('5'))
            Tab1Widget.paylabel.setText(str(paymanay))
            Tab1Widget.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_6:
            paymanay = (str(paymanay)+str('6'))
            Tab1Widget.paylabel.setText(str(paymanay))
            Tab1Widget.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_7:
            paymanay = (str(paymanay)+str('7'))
            Tab1Widget.paylabel.setText(str(paymanay))
            Tab1Widget.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_8:
            paymanay = (str(paymanay)+str('8'))
            Tab1Widget.paylabel.setText(str(paymanay))
            Tab1Widget.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_9:
            paymanay = (str(paymanay)+str('9'))
            Tab1Widget.paylabel.setText(str(paymanay))
            Tab1Widget.pay2label.setText(str(paymanay))


        if e.key() == Qt.Key_0:
            paymanay = (str(paymanay)+str('0'))
            Tab1Widget.paylabel.setText(str(paymanay))
            Tab1Widget.pay2label.setText(str(paymanay))

        if e.key() == Qt.Key_Control:
            allplice = (int(0))
            paymanay = (str(''))
            change = (int(0))
            Tab1Widget.paylabel.setText(str(paymanay))
            Tab1Widget.pay2label.setText(str(paymanay))
            Tab1Widget.plicelabel.setText(str(allplice))
            Tab1Widget.changelabel.setText(str(change))
            t = datetime.datetime.now()
            time = t.time()
            string = str("クリア")
            one = str("")
            display_list = [str(time), string, one]
            Tab1Widget.model.additeminrow(display_list)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":以下クリア:"+"￥0"+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()



        if e.key() == Qt.Key_Return:
            change = (int(paymanay)-int(allplice))
            Tab1Widget.changelabel.setText(str(change))
            t = datetime.datetime.now()
            time = t.time()
            string = str('=====総額=======')
            one = str(allplice)
            display_list = [str(time), string, one]
            Tab1Widget.model.additeminrow(display_list)

            datatime = datetime.datetime.now()
            datasorce = (str(datatime)+":===総額===:"+"￥"+str(allplice)+"("+zcode+")"+"\n")
            data = os.path.abspath("取引データ.txt")
            file = open(data, 'a')
            file.write(datasorce)
            file.close()





    def settingfun(self):

        global nametag1
        global nametag2
        global nametag3
        global nametag4
        global nametag5
        global nametag6
        global nametag7
        global nametag8
        global nametag9
        global nametag10
        global nametag11
        global nametag12
        global nametag13
        global nametag14
        global nametag15
        global nametag16
        global nametag17
        global nametag18
        global nametag19
        global nametag20
        global nametag21
        global nametag22
        global nametag23
        global nametag24
        global nametag25
        global nametag26
        global nametag27
        global nametag28
        global nametag29
        global nametag30
        global nametag31
        global nametag32
        global nametag33
        global nametag34
        global nametag35
        global nametag36




        global setpath
        global p1tag
        global p2tag
        global p3tag
        global p4tag
        global p5tag
        global p6tag
        global p7tag
        global p8tag
        global p9tag
        global p10tag
        global p11tag
        global p12tag
        global p13tag
        global p14tag
        global p15tag
        global p16tag
        global p17tag
        global p18tag
        global p18tag
        global p19tag
        global p20tag
        global p21tag
        global p22tag
        global p23tag
        global p24tag
        global p25tag
        global p26tag
        global p27tag
        global p28tag
        global p29tag
        global p30tag
        global p31tag
        global p32tag
        global p33tag
        global p34tag
        global p35tag
        global p36tag

        global snametag1
        global snametag2
        global snametag3
        global snametag4
        global snametag5
        global snametag6
        global snametag7
        global snametag8
        global snametag9
        global snametag10
        global snametag11
        global snametag12
        global snametag13
        global snametag14
        global snametag15
        global snametag16
        global snametag17
        global snametag18
        global snametag19
        global snametag20
        global snametag21
        global snametag22
        global snametag23
        global snametag24
        global snametag25
        global snametag26
        global snametag27
        global snametag28
        global snametag29
        global snametag30
        global snametag31
        global snametag32
        global snametag33
        global snametag34
        global snametag35
        global snametag36


        global sp1tag
        global sp2tag
        global sp3tag
        global sp4tag
        global sp5tag
        global sp6tag
        global sp7tag
        global sp8tag
        global sp9tag
        global sp10tag
        global sp11tag
        global sp12tag
        global sp13tag
        global sp14tag
        global sp15tag
        global sp16tag
        global sp17tag
        global sp18tag
        global sp18tag
        global sp19tag
        global sp20tag
        global sp21tag
        global sp22tag
        global sp23tag
        global sp24tag
        global sp25tag
        global sp26tag
        global sp27tag
        global sp28tag
        global sp29tag
        global sp30tag
        global sp31tag
        global sp32tag
        global sp33tag
        global sp34tag
        global sp35tag
        global sp36tag
        global info

        global dnametag1
        global dnametag2
        global dnametag3
        global dnametag4
        global dnametag5
        global dnametag6
        global dnametag7
        global dnametag8
        global dnametag9
        global dnametag10
        global dnametag11
        global dnametag12
        global dnametag13
        global dnametag14
        global dnametag15
        global dnametag16
        global dnametag17
        global dnametag18
        global dnametag19
        global dnametag20
        global dnametag21
        global dnametag22
        global dnametag23
        global dnametag24
        global dnametag25
        global dnametag26
        global dnametag27
        global dnametag28
        global dnametag29
        global dnametag30
        global dnametag31
        global dnametag32
        global dnametag33
        global dnametag34
        global dnametag35
        global dnametag36




        global setpath
        global dp1tag
        global dp2tag
        global dp3tag
        global dp4tag
        global dp5tag
        global dp6tag
        global dp7tag
        global dp8tag
        global dp9tag
        global dp10tag
        global dp11tag
        global dp12tag
        global dp13tag
        global dp14tag
        global dp15tag
        global dp16tag
        global dp17tag
        global dp18tag
        global dp18tag
        global dp19tag
        global dp20tag
        global dp21tag
        global dp22tag
        global dp23tag
        global dp24tag
        global dp25tag
        global dp26tag
        global dp27tag
        global dp28tag
        global dp29tag
        global dp30tag
        global dp31tag
        global dp32tag
        global dp33tag
        global dp34tag
        global dp35tag
        global dp36tag



        tag = openpyxl.load_workbook(setpath)
        act = tag.active
        #ここから商品名

        nametag1 = act["B3"].value
        nametag2 = act["B4"].value
        nametag3 = act["B5"].value
        nametag4 = act["B6"].value
        nametag5 = act["B7"].value
        nametag6 = act["B8"].value
        nametag7 = act["B9"].value
        nametag8 = act["B10"].value
        nametag9 = act["B11"].value
        nametag10 = act["B12"].value
        nametag11 = act["B13"].value
        nametag12 = act["B14"].value
        nametag13 = act["B15"].value
        nametag14 = act["B16"].value
        nametag15 = act["B17"].value
        nametag16 = act["B18"].value
        nametag17 = act["B19"].value
        nametag18 = act["B20"].value


        nametag19 = act["H3"].value
        nametag20 = act["H4"].value
        nametag21 = act["H5"].value
        nametag22 = act["H6"].value
        nametag23 = act["H7"].value
        nametag24 = act["H8"].value
        nametag25 = act["H9"].value
        nametag26 = act["H10"].value
        nametag27 = act["H11"].value
        nametag28 = act["H12"].value
        nametag29 = act["H13"].value
        nametag30 = act["H14"].value
        nametag31 = act["H15"].value
        nametag32 = act["H16"].value
        nametag33 = act["H17"].value
        nametag34 = act["H18"].value
        nametag35 = act["H19"].value
        nametag36 = act["H20"].value


        p1tag = act["C3"].value
        p2tag = act["C4"].value
        p3tag = act["C5"].value
        p4tag = act["C6"].value
        p5tag = act["C7"].value
        p6tag = act["C8"].value
        p7tag = act["C9"].value
        p8tag = act["C10"].value
        p9tag = act["C11"].value
        p10tag = act["C12"].value
        p11tag = act["C13"].value
        p12tag = act["C14"].value
        p13tag = act["C15"].value
        p14tag = act["C16"].value
        p15tag = act["C17"].value
        p16tag = act["C18"].value
        p17tag = act["C19"].value
        p18tag = act["C20"].value

        p19tag = act["I3"].value
        p20tag = act["I4"].value
        p21tag = act["I5"].value
        p22tag = act["I6"].value
        p23tag = act["I7"].value
        p24tag = act["I8"].value
        p25tag = act["I9"].value
        p26tag = act["I10"].value
        p27tag = act["I11"].value
        p28tag = act["I12"].value
        p29tag = act["I13"].value
        p30tag = act["I14"].value
        p31tag = act["I15"].value
        p32tag = act["I16"].value
        p33tag = act["I17"].value
        p34tag = act["I18"].value
        p35tag = act["I19"].value
        p36tag = act["I20"].value



        snametag1 = act["B27"].value
        snametag2 = act["B28"].value
        snametag3 = act["B29"].value
        snametag4 = act["B30"].value
        snametag5 = act["B31"].value
        snametag6 = act["B32"].value
        snametag7 = act["B33"].value
        snametag8 = act["B34"].value
        snametag9 = act["B35"].value
        snametag10 = act["B36"].value
        snametag11 = act["B37"].value
        snametag12 = act["B38"].value
        snametag13 = act["B39"].value
        snametag14 = act["B40"].value
        snametag15 = act["B41"].value
        snametag16 = act["B42"].value
        snametag17 = act["B43"].value
        snametag18 = act["B44"].value

        snametag19 = act["H27"].value
        snametag20 = act["H28"].value
        snametag21 = act["H29"].value
        snametag22 = act["H30"].value
        snametag23 = act["H31"].value
        snametag24 = act["H32"].value
        snametag25 = act["H33"].value
        snametag26 = act["H34"].value
        snametag27 = act["H35"].value
        snametag28 = act["H36"].value
        snametag29 = act["H37"].value
        snametag30 = act["H38"].value
        snametag31 = act["H39"].value
        snametag32 = act["H40"].value
        snametag33 = act["H41"].value
        snametag34 = act["H42"].value
        snametag35 = act["H43"].value
        snametag36 = act["H44"].value


        sp1tag = act["C27"].value
        sp2tag = act["C28"].value
        sp3tag = act["C29"].value
        sp4tag = act["C30"].value
        sp5tag = act["C31"].value
        sp6tag = act["C32"].value
        sp7tag = act["C33"].value
        sp8tag = act["C34"].value
        sp9tag = act["C35"].value
        sp10tag = act["C36"].value
        sp11tag = act["C37"].value
        sp12tag = act["C38"].value
        sp13tag = act["C39"].value
        sp14tag = act["C40"].value
        sp15tag = act["C41"].value
        sp16tag = act["C42"].value
        sp17tag = act["C43"].value
        sp18tag = act["C44"].value

        sp19tag = act["I27"].value
        sp20tag = act["I28"].value
        sp21tag = act["I29"].value
        sp22tag = act["I30"].value
        sp23tag = act["I31"].value
        sp24tag = act["I32"].value
        sp25tag = act["I33"].value
        sp26tag = act["I34"].value
        sp27tag = act["I35"].value
        sp28tag = act["I36"].value
        sp29tag = act["I37"].value
        sp30tag = act["I38"].value
        sp31tag = act["I39"].value
        sp32tag = act["I40"].value
        sp33tag = act["I41"].value
        sp34tag = act["I42"].value
        sp35tag = act["I43"].value
        sp36tag = act["I44"].value

        dnametag1 = act["B50"].value
        dnametag2 = act["B51"].value
        dnametag3 = act["B52"].value
        dnametag4 = act["B53"].value
        dnametag5 = act["B54"].value
        dnametag6 = act["B55"].value
        dnametag7 = act["B56"].value
        dnametag8 = act["B57"].value
        dnametag9 = act["B58"].value
        dnametag10 = act["B59"].value
        dnametag11 = act["B60"].value
        dnametag12 = act["B61"].value
        dnametag13 = act["B62"].value
        dnametag14 = act["B63"].value
        dnametag15 = act["B64"].value
        dnametag16 = act["B65"].value
        dnametag17 = act["B66"].value
        dnametag18 = act["B67"].value


        dnametag19 = act["H50"].value
        dnametag20 = act["H51"].value
        dnametag21 = act["H52"].value
        dnametag22 = act["H53"].value
        dnametag23 = act["H54"].value
        dnametag24 = act["H55"].value
        dnametag25 = act["H56"].value
        dnametag26 = act["H57"].value
        dnametag27 = act["H58"].value
        dnametag28 = act["H59"].value
        dnametag29 = act["H60"].value
        dnametag30 = act["H61"].value
        dnametag31 = act["H62"].value
        dnametag32 = act["H63"].value
        dnametag33 = act["H64"].value
        dnametag34 = act["H65"].value
        dnametag35 = act["H66"].value
        dnametag36 = act["H67"].value


        dp1tag = act["C50"].value
        dp2tag = act["C51"].value
        dp3tag = act["C52"].value
        dp4tag = act["C53"].value
        dp5tag = act["C54"].value
        dp6tag = act["C55"].value
        dp7tag = act["C56"].value
        dp8tag = act["C57"].value
        dp9tag = act["C58"].value
        dp10tag = act["C59"].value
        dp11tag = act["C60"].value
        dp12tag = act["C61"].value
        dp13tag = act["C62"].value
        dp14tag = act["C63"].value
        dp15tag = act["C64"].value
        dp16tag = act["C65"].value
        dp17tag = act["C66"].value
        dp18tag = act["C67"].value

        dp19tag = act["I50"].value
        dp20tag = act["I51"].value
        dp21tag = act["I52"].value
        dp22tag = act["I53"].value
        dp23tag = act["I54"].value
        dp24tag = act["I55"].value
        dp25tag = act["I56"].value
        dp26tag = act["I57"].value
        dp27tag = act["I58"].value
        dp28tag = act["I59"].value
        dp29tag = act["I60"].value
        dp30tag = act["I61"].value
        dp31tag = act["I62"].value
        dp32tag = act["I63"].value
        dp33tag = act["I64"].value
        dp34tag = act["I65"].value
        dp35tag = act["I66"].value
        dp36tag = act["I67"].value



        info = act["F21"].value

        Tab1Widget.name1.setText(str(nametag1))
        Tab1Widget.name2.setText(str(nametag2))
        Tab1Widget.name3.setText(str(nametag3))
        Tab1Widget.name4.setText(str(nametag4))
        Tab1Widget.name5.setText(str(nametag5))
        Tab1Widget.name6.setText(str(nametag6))
        Tab1Widget.name7.setText(str(nametag7))
        Tab1Widget.name8.setText(str(nametag8))
        Tab1Widget.name9.setText(str(nametag9))
        Tab1Widget.name10.setText(str(nametag10))
        Tab1Widget.name11.setText(str(nametag11))
        Tab1Widget.name12.setText(str(nametag12))
        Tab1Widget.name13.setText(str(nametag13))
        Tab1Widget.name14.setText(str(nametag14))
        Tab1Widget.name15.setText(str(nametag15))
        Tab1Widget.name16.setText(str(nametag16))
        Tab1Widget.name17.setText(str(nametag17))
        Tab1Widget.name18.setText(str(nametag18))

        Tab1Widget.name19.setText(str(nametag19))
        Tab1Widget.name20.setText(str(nametag20))
        Tab1Widget.name21.setText(str(nametag21))
        Tab1Widget.name22.setText(str(nametag22))
        Tab1Widget.name23.setText(str(nametag23))
        Tab1Widget.name24.setText(str(nametag24))
        Tab1Widget.name25.setText(str(nametag25))
        Tab1Widget.name26.setText(str(nametag26))
        Tab1Widget.name27.setText(str(nametag27))
        Tab1Widget.name28.setText(str(nametag28))
        Tab1Widget.name29.setText(str(nametag29))
        Tab1Widget.name30.setText(str(nametag30))
        Tab1Widget.name31.setText(str(nametag31))
        Tab1Widget.name32.setText(str(nametag32))
        Tab1Widget.name33.setText(str(nametag33))
        Tab1Widget.name34.setText(str(nametag34))
        Tab1Widget.name35.setText(str(nametag35))
        Tab1Widget.name36.setText(str(nametag36))

        Tab1Widget.sname1.setText(str(snametag1))
        Tab1Widget.sname2.setText(str(snametag2))
        Tab1Widget.sname3.setText(str(snametag3))
        Tab1Widget.sname4.setText(str(snametag4))
        Tab1Widget.sname5.setText(str(snametag5))
        Tab1Widget.sname6.setText(str(snametag6))
        Tab1Widget.sname7.setText(str(snametag7))
        Tab1Widget.sname8.setText(str(snametag8))
        Tab1Widget.sname9.setText(str(snametag9))
        Tab1Widget.sname10.setText(str(snametag10))
        Tab1Widget.sname11.setText(str(snametag11))
        Tab1Widget.sname12.setText(str(snametag12))
        Tab1Widget.sname13.setText(str(snametag13))
        Tab1Widget.sname14.setText(str(snametag14))
        Tab1Widget.sname15.setText(str(snametag15))
        Tab1Widget.sname16.setText(str(snametag16))
        Tab1Widget.sname17.setText(str(snametag17))
        Tab1Widget.sname18.setText(str(snametag18))

        Tab1Widget.sname19.setText(str(snametag19))
        Tab1Widget.sname20.setText(str(snametag20))
        Tab1Widget.sname21.setText(str(snametag21))
        Tab1Widget.sname22.setText(str(snametag22))
        Tab1Widget.sname23.setText(str(snametag23))
        Tab1Widget.sname24.setText(str(snametag24))
        Tab1Widget.sname25.setText(str(snametag25))
        Tab1Widget.sname26.setText(str(snametag26))
        Tab1Widget.sname27.setText(str(snametag27))
        Tab1Widget.sname28.setText(str(snametag28))
        Tab1Widget.sname29.setText(str(snametag29))
        Tab1Widget.sname30.setText(str(snametag30))
        Tab1Widget.sname31.setText(str(snametag31))
        Tab1Widget.sname32.setText(str(snametag32))
        Tab1Widget.sname33.setText(str(snametag33))
        Tab1Widget.sname34.setText(str(snametag34))
        Tab1Widget.sname35.setText(str(snametag35))
        Tab1Widget.sname36.setText(str(snametag36))

        Tab1Widget.dname1.setText(str(dnametag1))
        Tab1Widget.dname2.setText(str(dnametag2))
        Tab1Widget.dname3.setText(str(dnametag3))
        Tab1Widget.dname4.setText(str(dnametag4))
        Tab1Widget.dname5.setText(str(dnametag5))
        Tab1Widget.dname6.setText(str(dnametag6))
        Tab1Widget.dname7.setText(str(dnametag7))
        Tab1Widget.dname8.setText(str(dnametag8))
        Tab1Widget.dname9.setText(str(dnametag9))
        Tab1Widget.dname10.setText(str(dnametag10))
        Tab1Widget.dname11.setText(str(dnametag11))
        Tab1Widget.dname12.setText(str(dnametag12))
        Tab1Widget.dname13.setText(str(dnametag13))
        Tab1Widget.dname14.setText(str(dnametag14))
        Tab1Widget.dname15.setText(str(dnametag15))
        Tab1Widget.dname16.setText(str(dnametag16))
        Tab1Widget.dname17.setText(str(dnametag17))
        Tab1Widget.dname18.setText(str(dnametag18))

        Tab1Widget.dname19.setText(str(dnametag19))
        Tab1Widget.dname20.setText(str(dnametag20))
        Tab1Widget.dname21.setText(str(dnametag21))
        Tab1Widget.dname22.setText(str(dnametag22))
        Tab1Widget.dname23.setText(str(dnametag23))
        Tab1Widget.dname24.setText(str(dnametag24))
        Tab1Widget.dname25.setText(str(dnametag25))
        Tab1Widget.dname26.setText(str(dnametag26))
        Tab1Widget.dname27.setText(str(dnametag27))
        Tab1Widget.dname28.setText(str(dnametag28))
        Tab1Widget.dname29.setText(str(dnametag29))
        Tab1Widget.dname30.setText(str(dnametag30))
        Tab1Widget.dname31.setText(str(dnametag31))
        Tab1Widget.dname32.setText(str(dnametag32))
        Tab1Widget.dname33.setText(str(dnametag33))
        Tab1Widget.dname34.setText(str(dnametag34))
        Tab1Widget.dname35.setText(str(dnametag35))
        Tab1Widget.dname36.setText(str(dnametag36))





































app = QApplication(sys.argv)
main_window = MainWin()
main_window.show()
main_window.raise_()
sys.exit(app.exec_())
main()
